from datetime import datetime, date, timedelta
from datetime import time as tm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side, Color, Protection
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.styles import numbers
from openpyxl.utils.datetime import CALENDAR_MAC_1904
import requests
from time import sleep
import os
import copy     # Issue #4
try:
    import xlrd4 as xlrd
except Exception:
    import xlrd

RETRIES=6

class XLS2XLSX:
    """Convert an xls file into an xlsx file.  Everything is supported except for the things
       not supported by xlrd, which include:
        1. Conditional Formatting
        2. Formulas
        3. Charts and Images
        4. Pivot Tables
       If this xls file is in html format, then we call HTMLXLS2XLSX to convert it.
        """

    def __init__(self, f, dirname='.', ignore_workbook_corruption=False):
        """f is a url, filename, file object, or xls file contents as a bytes object"""
        self.dirname = dirname
        self.ignore_workbook_corruption = ignore_workbook_corruption
        if isinstance(f, bytes):
            self.contents = f
        else:
            if isinstance(f, str):
                self.dirname = os.path.split(f)[0]
            self.contents = XLS2XLSX.read(f)

        self.h2x = None
        try:
            self.book = xlrd.open_workbook(file_contents=self.contents, formatting_info=True,
                                           ignore_workbook_corruption=ignore_workbook_corruption)
            self.date_mode = self.book.datemode
        except xlrd.XLRDError as e:
            if str(e).startswith('Unsupported format'):
                from .htmlxls2xlsx import HTMLXLS2XLSX
                self.h2x = HTMLXLS2XLSX(self.contents, self.dirname)
            else:
                raise ValueError(e)     # Issue #3
        except xlrd.compdoc.CompDocError:
            try:
                self.book = xlrd.open_workbook(file_contents=self.contents, formatting_info=True,
                                           ignore_workbook_corruption=True)
                self.date_mode = self.book.datemode
            except Exception:
                self.book = None        # completely ignore corruption that cannot be corrected

    @staticmethod
    def read(f, retries=RETRIES):
        """Read from either a URL or a filename or file-like object."""
        if isinstance(f, str):
            if '://' in f:  # URL
                for r in range(retries):
                    try:
                        resp = requests.get(f)
                        resp.raise_for_status()
                        return resp.content
                    except Exception:
                        if r == retries-1:
                            raise
                        sleep(2)
            with open(f, 'rb') as t:
                return t.read()
        else:
            return f.read()

    def xls_date_to_xlsx(self, value):
        date_tuple = xlrd.xldate_as_tuple(value, self.date_mode)
        if date_tuple == (0, 0, 0, 0, 0, 0):
            return datetime(1900, 1, 1, 0, 0, 0)
        elif date_tuple[0:3] == (0, 0, 0):
            return tm(date_tuple[3], date_tuple[4], date_tuple[5])
        elif date_tuple[3:6] == (0, 0, 0):
            return date(date_tuple[0], date_tuple[1], date_tuple[2])
        return datetime(date_tuple[0], date_tuple[1], date_tuple[2], date_tuple[3], date_tuple[4], date_tuple[5])

    def xls_color_to_xlsx(self, color_ndx):
        black = (0, 0, 0)
        color_tuple = self.book.colour_map.get(color_ndx, black)
        if color_tuple is None:
            color_tuple = black
        return Color(f'{color_tuple[0]:02X}{color_tuple[1]:02X}{color_tuple[2]:02X}')

    def xls_width_to_xlsx(self, width):
        return width / 256      # to chars

    def xls_height_to_xlsx(self, height):
        return height / 20      # to pts

    def xls_style_to_xlsx(self, xf_ndx):
        """Convert an xls xf_ndx into a 6-tuple of styles for xlsx"""
        font = Font()
        fill = PatternFill()
        border = Border()
        alignment = Alignment()
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        if xf_ndx < len(self.book.xf_list):
            xf = self.book.xf_list[xf_ndx]

            try:            # Avoidance for issue #11 (though I cannot duplicate the problem w/o the input file)
                xls_font = self.book.font_list[xf.font_index]       # Font object
                font.b = xls_font.bold
                font.i = xls_font.italic
                if xls_font.character_set:
                    font.charset = xls_font.character_set
                font.color = self.xls_color_to_xlsx(xls_font.colour_index)
                escapement = xls_font.escapement        # 1=Superscript, 2=Subscript
                family = xls_font.family                # FIXME: 0=Any, 1=Roman, 2=Sans, 3=monospace, 4=Script, 5=Old English/Franktur
                font.name = xls_font.name
                font.sz = self.xls_height_to_xlsx(xls_font.height)    # A twip = 1/20 of a pt
                if xls_font.struck_out:
                    font.strike = xls_font.struck_out
                if xls_font.underline_type:
                    font.u = ('single', 'double')[(xls_font.underline_type&3)-1]
            except Exception:
                pass

            xls_format = self.book.format_map[xf.format_key]    # Format object
            number_format = xls_format.format_str

            if False:               # xlrd says all cells are locked even if the sheet isn't protected!
                protection.locked = xf.protection.cell_locked
            protection.hidden = xf.protection.formula_hidden

            fill_patterns = {0x00:'none', 0x01:'solid', 0x02:'mediumGray', 0x03:'darkGray', 0x04:'lightGray', 
                    0x05:'darkHorizontal', 0x06:'darkVertical', 0x07:'darkDown', 0x08:'darkUp', 0x09:'darkGrid', 
                    0x0A:'darkTrellis', 0x0B:'lightHorizontal', 0x0C:'lightVertical', 0x0D:'lightDown', 0x0E:'lightUp',
                    0x0F:'lightGrid', 0x10:'lightTrellis', 0x11:'gray125', 0x12:'gray0625'
                    }
            fill_pattern = xf.background.fill_pattern
            fill_background_color = self.xls_color_to_xlsx(xf.background.background_colour_index)
            fill_pattern_color = self.xls_color_to_xlsx(xf.background.pattern_colour_index)
            fill.patternType = fill_patterns.get(fill_pattern, 'none')
            fill.bgColor = fill_background_color
            fill.fgColor = fill_pattern_color

            horizontal = {0:'general', 1:'left', 2:'center', 3:'right', 4:'fill', 5:'justify', 6:'centerContinuous', 7:'distributed'}
            vertical = {0:'top', 1:'center', 2:'bottom', 3:'justify', 4:'distributed'}
            hor_align = horizontal.get(xf.alignment.hor_align, None)
            if hor_align:
                alignment.horizontal = hor_align
            vert_align = vertical.get(xf.alignment.vert_align, None)
            if vert_align:
                alignment.vertical = vert_align
            alignment.textRotation = xf.alignment.rotation
            alignment.wrap_text = xf.alignment.text_wrapped
            alignment.indent = xf.alignment.indent_level
            alignment.shrink_to_fit = xf.alignment.shrink_to_fit

            border_styles = {0: None, 1:'thin', 2:'medium', 3:'dashed', 4:'dotted',
                5:'thick', 6:'double', 7:'hair', 8:'mediumDashed', 9:'dashDot',
                10:'mediumDashDot', 11:'dashDotDot', 12:'mediumDashDotDot',
                13:'slantDashDot',}
            xls_border = xf.border
            top = Side(style=border_styles.get(xls_border.top_line_style), color=self.xls_color_to_xlsx(xls_border.top_colour_index))
            bottom = Side(style=border_styles.get(xls_border.bottom_line_style), color=self.xls_color_to_xlsx(xls_border.bottom_colour_index))
            left = Side(style=border_styles.get(xls_border.left_line_style), color=self.xls_color_to_xlsx(xls_border.left_colour_index))
            right = Side(style=border_styles.get(xls_border.right_line_style), color=self.xls_color_to_xlsx(xls_border.right_colour_index))
            diag = Side(style=border_styles.get(xls_border.diag_line_style), color=self.xls_color_to_xlsx(xls_border.diag_colour_index))
            border.top = top
            border.bottom = bottom
            border.left = left
            border.right = right
            border.diagonal = diag
            border.diagonalDown = xls_border.diag_down
            border.diagonalUp = xls_border.diag_up

        return (font, fill, border, alignment, number_format, protection)

    def to_xlsx(self, filename=None):
        """Convert to xlsx using openpyxl.  If filename is not None, then the result
        is written to that file, and the filename is returned, else the workbook is returned.
        """

        if self.h2x:
            return self.h2x.to_xlsx(filename=filename)

        if self.book is None and self.ignore_workbook_corruption:
            return None     # Couldn't be loaded - nothing we can do

        wb = Workbook()         # creates one worksheet
        ws = wb.active

        wb.properties.lastModifiedBy = self.book.user_name
        if self.date_mode:
            wb.epoch = CALENDAR_MAC_1904

        for sheet in self.book.sheets():
            if ws:
                ws.title = sheet.name
            else:
                ws = wb.create_sheet(sheet.name)

            for col_ndx, info in sheet.colinfo_map.items():
                col = get_column_letter(col_ndx+1)
                if info.hidden:
                    ws.column_dimensions[col].hidden = True
                else:
                    ws.column_dimensions[col].width = self.xls_width_to_xlsx(info.width)
            for row_ndx, info in sheet.rowinfo_map.items():
                row = row_ndx+1
                if info.hidden:
                    ws.row_dimensions[row].hidden = True
                else:
                    ws.row_dimensions[row].height = self.xls_height_to_xlsx(info.height)
                ws.row_dimensions[row].thickTop = info.additional_space_above
                ws.row_dimensions[row].thickBot = info.additional_space_below

            rows = sheet.nrows
            columns = sheet.ncols

            for row in range(rows):
                for col in range(columns):
                    cell_type = sheet.cell_type(row, col)
                    value = sheet.cell_value(row, col)
                    if cell_type == xlrd.XL_CELL_DATE:
                        try:        # Issue #5: Just keep 'bad' dates as float numbers
                            value = self.xls_date_to_xlsx(value)
                        except Exception:
                            pass
                    elif cell_type == xlrd.XL_CELL_NUMBER:
                        try:
                            ival = int(value)
                            if ival == value:
                                value = ival
                        except Exception:
                            pass
                    elif cell_type == xlrd.XL_CELL_ERROR:
                        if value in xlrd.biffh.error_text_from_code:
                            value = xlrd.biffh.error_text_from_code[value]
                        else:
                            value = '#N/A'
                    elif cell_type in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        value = ''
                    elif cell_type == xlrd.XL_CELL_BOOLEAN:
                        value = ('false', 'true')[value]

                    rw = row+1
                    cc = col+1
                    ws.cell(rw, cc).value = value
                    font, fill, border, alignment, number_format, protection = self.xls_style_to_xlsx(sheet.cell_xf_index(row, col))
                    #if number_format != 'General':
                        #print(f'({rw},{cc}).number_format = {number_format}')
                    if isinstance(value, str):
                        if '\n' in value and not alignment.wrap_text:
                            alignment = copy.deepcopy(alignment)
                            alignment.wrap_text = True
                        if value[-1:] == '%' and number_format == 'General':
                            number_format = numbers.FORMAT_PERCENTAGE
                    elif isinstance(value, datetime):
                        if number_format == 'General':
                            number_format = r'm\/d\/yyyy h\:mm\:ss AM/PM'
                        elif '/d/yy ' in number_format:
                            # For some unknown reason, 2-digit year formats are being stored
                            # when the user is specifying a 4-digit year
                            number_format = number_format.replace('/d/yy ', '/dd/yyyy ')
                        elif '/d\\/yy ' in number_format:
                            # For some unknown reason, 2-digit year formats are being stored
                            # when the user is specifying a 4-digit year
                            number_format = number_format.replace('/d\\/yy ', '/dd\\/yyyy ')
                        elif '/yy ' in number_format:
                            number_format = number_format.replace('/yy ', '/yyyy ')
                    elif isinstance(value, date):
                        if number_format == 'General':
                            number_format = r'm\/d\/yyyy'
                        elif number_format.endswith('/d/yy'):
                            number_format = number_format.replace('/d/yy', '/dd/yyyy')
                        elif number_format.endswith('/d\\/yy'):
                            number_format = number_format.replace('/d\\/yy', '/dd\\/yyyy')
                        elif number_format.endswith('/yy'):
                            number_format += 'yy'       # Change to yyyy
                    elif isinstance(value, tm):
                        if number_format == 'General':
                            number_format = r'h\:mm\:ss AM/PM'
                    elif isinstance(value, timedelta):
                        if number_format == 'General':
                            number_format = '[h]:mm:ss'
                    ws.cell(rw, cc).font = font
                    ws.cell(rw, cc).fill = fill
                    ws.cell(rw, cc).border = border
                    ws.cell(rw, cc).alignment = alignment
                    #if number_format != 'General':
                        #print(f'({rw},{cc}).number_format = {number_format}, .value = {value}, type = {type(value)}')
                    ws.cell(rw, cc).number_format = number_format
                    ws.cell(rw, cc).protection = protection
                    if protection.locked or protection.hidden:
                        ws.protection.sheet = True
                    tup = (row, col)
                    if tup in sheet.hyperlink_map:
                        hyperlink = sheet.hyperlink_map[tup].url_or_path
                        ws.cell(rw, cc).hyperlink = hyperlink
                        #ws.cell(rw, cc).style = 'Hyperlink'
                    if tup in sheet.cell_note_map:
                        comment = sheet.cell_note_map[tup]
                        ws.cell(rw, cc).comment = Comment(comment.text, comment.author)
                    image = False       # FIXME (after fixing xlrd)
                    if image:
                        image.anchor = f'{cl}{rw}'
                        ws.add_image(image)

            for crange in sheet.merged_cells:
                rlo, rhi, clo, chi = crange
                ws.merge_cells(start_row=rlo+1,
                        start_column=clo+1,
                        end_row=rhi,
                        end_column=chi)

            if sheet.visibility:
                ws.sheet_state = 'hidden'

            if sheet.vert_split_pos != 0 or sheet.horz_split_pos != 0:
                row = sheet.horz_split_pos + 1
                col = sheet.vert_split_pos + 1
                ws.freeze_panes = f'{get_column_letter(col)}{row}'

            ws = None

        if filename:
            wb.save(filename=filename)
            return filename
        return wb
