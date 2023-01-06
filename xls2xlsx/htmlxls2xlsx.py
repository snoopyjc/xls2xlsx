from bs4 import BeautifulSoup, UnicodeDammit, NavigableString, Comment, CData, ProcessingInstruction, Declaration, Doctype    # pip install beautifulsoup4
import quopri
from bs4 import GuessedAtParserWarning
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side, Color
from openpyxl.comments import Comment as OpenpyxlComment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.styles import numbers
from openpyxl.styles.numbers import BUILTIN_FORMATS, BUILTIN_FORMATS_REVERSE
import requests
import copy
import io
import os
import re
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import cssutils
import webcolors
import traceback
import logging
import email            # For mht files
import shutil
from functools import lru_cache
from dateutil.parser import parse as date_parse
from datetime import datetime, date, timedelta
from datetime import time as tm
from urllib.parse import urljoin
from time import sleep
from PIL import ImageFont
import math
from fontTools import ttLib          # pip install fonttools
import yaml
import sys
import calendar
import currency_symbols._constants as currency_symbols_constants

TRACE=False

cssutils.log.setLevel(logging.CRITICAL) # Remove 'Unknown Property name' messages
BUILTIN_FORMATS[14] = 'm/d/yyyy'    # See https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1534
del BUILTIN_FORMATS_REVERSE['mm-dd-yy']
BUILTIN_FORMATS_REVERSE['m/d/yyyy'] = 14
REQUESTS_HEADER = {'User-Agent': 'xls2xlsx/0.2.0 (https://pypi.org/project/xls2xlsx/; snoopyjc@gmail.com) htmlxls2xlsx.py/0.2.0'}

class FontUtils:
    LINE_HEIGHT_FACTOR = 1.25        # line_height in px = Font size in px * LINE_HEIGHT_FACTOR
    def __init__(self):
        self.skinny_chars = "1!|iIl.,;:' "
        self.upper_chars = "ABCDEFGHJKLMNOPQRSTUVXYZmw()[]$&*-+{}<>/?"
        self.fat_chars = 'MW@#%_'
        fontnames_file = os.path.join(os.path.dirname(__file__), 'fontnames.yaml')
        self.name_to_path = {}
        if os.path.isfile(fontnames_file):
            with open(fontnames_file, 'r') as fn:
                self.name_to_path = yaml.safe_load(fn)
        else:
            dirs = ['fonts']
            # Code 'borrowed' from PIL:
            if sys.platform == "win32":
                windir = os.environ.get("WINDIR")
                if windir:
                    dirs.append(os.path.join(windir, "fonts"))
            elif sys.platform in ("linux", "linux2"):
                lindirs = os.environ.get("XDG_DATA_DIRS", "")
                if not lindirs:
                    lindirs = "/usr/share"
                dirs += [os.path.join(lindir, "fonts") for lindir in lindirs.split(":")]
            elif sys.platform == "darwin":
                dirs += ["/Library/Fonts", "/System/Library/Fonts",
                    os.path.expanduser("~/Library/Fonts"),
                ]
            FONT_SPECIFIER_NAME_ID = 4
            for d in dirs:
                if os.path.isdir(d):
                    files = os.listdir(d)
                    for f in files:
                        path = os.path.join(d, f)
                        try:
                            tt = ttLib.TTFont(path)
                            name = None
                            for r in tt['name'].names:
                                if r.nameID == FONT_SPECIFIER_NAME_ID:
                                    if b'\000' in r.string:
                                        try:
                                            name = str(r.string, 'utf-16-be')
                                            break
                                        except Exception:
                                            pass
                                    else:
                                        name = str(r.string, 'latin-1')
                                        break
                            if name:
                                nl = name.lower()
                                if nl != name:
                                    self.name_to_path[nl] = '!' + name   # '!' means alias
                                self.name_to_path[name] = path
                            else:
                                print(f'No name for {path}')
                        except Exception:
                            pass
            with open(fontnames_file, 'w') as fn:
                yaml.dump(self.name_to_path, fn)

    def get_real_font_name(self, name, bold=False, italic=False):
        """Given a font name which may not be in the proper case, return the real font name, or None if not found"""
        fn = name.lower()
        if bold:
            fn += ' bold'
        if italic:
            fn += ' italic'
        result = self.name_to_path.get(fn)
        if result is not None:
            if result[0] == '!':        # Alias
                return result[1:]       # Real name
            else:
                return fn               # We found it in lower case

    def get_font_path(self, name, bold=False, italic=False):
        """Given a font name (which may not be in the proper case), return the path to the font file or None if not found"""
        real_name = self.get_real_font_name(name, bold, italic)
        if real_name is None:
            return None
        return self.name_to_path.get(real_name)

    @lru_cache(maxsize=32)
    def get_font(self, name, size=10, bold=False, italic=False):
        # Pillow creates images that are 72 pixels per inch by default, and a point is 1/72 of an inch, so
        # we must convert the size to pixels in order to get the proper size.
        size = self.pt_to_px(size)
        size = math.ceil(size)
        font_path = self.get_font_path(name, bold, italic)
        if not font_path:
            return None
        return ImageFont.truetype(font_path, size)

    @staticmethod
    def str_to_filename(s, ext):    # pragma nocover
        """Convert this string to a valid filename (used for debugging only)"""
        result = re.sub(r'[.<>:"/\\|?*\[\]\s]', '_', s) + ext
        if len(result) > 100:
            result = result[:100] + ext
        return result

    def get_font_size(self, font, s):
        """Get the width and height of text 's' in the given font. 's' is a single line of text (without newlines)"""
        height = self.pt_to_px(font.sz) * self.LINE_HEIGHT_FACTOR
        if not font or not s:
            return (0, height)

        tt_font = self.get_font(font.name, font.sz, font.b, font.i)
        if tt_font:
            #return font.getsize_multiline(s, spacing=font.size/3)
            width, _ = tt_font.getsize(s)
            #width = self.pt_to_px(width)
            if TRACE:
                os.makedirs('trace', exist_ok=True)
                from PIL import ImageDraw
                from PIL import Image as PILImage
                img = PILImage.new('RGB', (math.ceil(width), math.ceil(height)), color='white')
                draw = ImageDraw.Draw(img)
                draw.text((0, 0), s, fill='black', font=tt_font)
                img.save(os.path.join('trace', self.str_to_filename(f'{font.name}_{font.sz}_{s}', '.png')))

        else:
            # Estimate the font size if we can't find the true answer
            width = 0
            for c in s:
                if c in self.skinny_chars:
                    width += 0.6
                elif c in self.fat_chars:
                    if font.b and font.name == 'Calibri':
                        width += 2.1
                    else:
                        width += 1.9
                elif c in self.upper_chars:
                    width += 1.4
                else:
                    width += 1
            if font.b and font.name != 'Calibri':
                width *= 1.1       # 10% wider for non-Calibri fonts in bold (like Arial)
            width += 1         # Give it some margin
            width *= (font.sz/11)
            width *= 7      # Convert chars to px
        return (width, height)

    def lines_needed(self, img_width, s, font):
        """How many lines are needed to render this text 's' using img_width pixels?"""
        if '\n' in s:
            lines = s.split('\n')
            result = 0
            for line in lines:
                result += self.lines_needed(img_width, line, font)
            if TRACE:
                print(f'lines_needed({img_width}, {s}, {font.sz}) = {result} (1)')
            return result
        number_of_lines = 0
        # count how many lines are needed to break the string into multi-lines that fit img_width
        line = ''
        for token in s.split():
            if line:
                line += ' ' + token
            else:
                line = token
            w = self.get_font_size(font, line)[0]
            if w > img_width:
                number_of_lines += int(w // img_width)
                line = token
            elif w == img_width:
                number_of_lines += 1
                line = ''
        if line:
            number_of_lines += 1

        if number_of_lines == 0:    # E.g. if you send the empty string
            number_of_lines = 1
        
        if TRACE:
            print(f'lines_needed({img_width}, {s}, {font.sz}) = {number_of_lines} (2)')
        return number_of_lines

    @staticmethod
    def pt_to_px(pt):
        if pt is None:
            return pt
        return pt / 0.75

    @staticmethod
    def px_to_pt(px):
        if px is None:
            return px
        return px * 0.75

class CSSStyle:
    RETRIES=6
    DEFAULT_POINT_SIZE = 10
    SPREADSHEET_WIDTH_PX = 1900      # e.g. to determine what width="5%" means
    SPREADSHEET_HEIGHT_PX = 800
    MAX_CELL_HEIGHT_PT = 409         # Excel limit on cell height
    MAX_CELL_WIDTH_UNITS = 255       # Excel limit on cell width
    MIN_CELL_HEIGHT_PT = FontUtils.px_to_pt(15)
    MIN_CELL_WIDTH_PX = 15
    DEFAULT_CELL_WIDTH_PX = 64

    def __init__(self):
        self.stylemap = {}
        self.pt_list = ('7.5pt', '10pt', '12pt', '13.5pt', '18pt', '24pt', '36pt')
        self.inherited_properties = { 'border-collapse', 'border-spacing', 'caption-side', 'color', 'cursor',
          'direction', 'empty-cells', 'font-family', 'font-size', 'font-style', 'font-variant', 'font-weight',
          'font-size-adjust', 'font-stretch', 'font', 'letter-spacing', 'line-height', 'list-style-image',
          'list-style-position', 'list-style-type', 'list-style', 'orphans', 'quotes', 'tab-size', 'text-align',
          'text-align-last', 'text-decoration-color', 'text-indent', 'text-justify', 'text-shadow', 'text-transform',
          'visibility', 'white-space', 'widows', 'word-break', 'word-spacing', 'word-wrap' }
        # We use Microsoft's special mso-style-parent to inherit the style of the <td> on the next element, which
        # we send thru depending on what tag it is.  This is used to style the entire cell, since we don't support
        # a per-element style (openpyxl doesn't support rich text)
        self.default_styles = """
.htmlxls2xlsx {background: inherit; background-color: inherit; 
    border: inherit; border-color: inherit; border-width: inherit; border-bottom-color: inherit; border-left-color: inherit;
    border-right-color: inherit; border-top-color: inherit; border-top: inherit; border-right: inherit; border-bottom: inherit; border-left: inherit;
    border-top-width: inherit; border-right-width: inherit; border-bottom-width: inherit; border-left-width: inherit;
    border-top-style: inherit; border-right-style: inherit; border-bottom-style: inherit; border-left-style: inherit;
    border-top-color: inherit; border-right-color: inherit; border-bottom-color: inherit; border-left-color: inherit;
    height: inherit; layout-flow: inherit; max-height: inherit; max-width: inherit; min-height: inherit; min-width: inherit; 
    mso-ignore: inherit; mso-char-indent-count: inherit; mso-number-format: inherit; mso-rotate: inherit; mso-text-control: inherit; 
    padding: inherit; padding-top: inherit; padding-right: inherit; padding-bottom: inherit; padding-left: inherit;
    text-decoration: inherit; vertical-align: inherit; width: inherit; writing-mode: inherit; }
.msocomtxt {display: none; }
.msocomanch {display: none; }
.msocomhide {display: none; }
a {mso-style-parent:htmlxls2xlsx; color: #0563C1; text-decoration: underline;}
b {mso-style-parent:htmlxls2xlsx; font-weight: bold;}
big {mso-style-parent:htmlxls2xlsx; font-size: 1.33em;}
center {mso-style-parent:htmlxls2xlsx; text-align: center;}
code {mso-style-parent:htmlxls2xlsx; font-family: monospace;}
div {mso-style-parent:htmlxls2xlsx;}
em {mso-style-parent:htmlxls2xlsx; font-style: italic;}
font {mso-style-parent:htmlxls2xlsx; }
h1 {mso-style-parent:htmlxls2xlsx; display: block; font-size: 2em; margin-top: 0.67em; margin-bottom: 0.67em; margin-left: 0; margin-right: 0; font-weight: bold;}
h2 {mso-style-parent:htmlxls2xlsx; display: block; font-size: 1.5em; margin-top: 0.83em; margin-bottom: 0.83em; margin-left: 0; margin-right: 0; font-weight: bold;}
h3 {mso-style-parent:htmlxls2xlsx; display: block; font-size: 1.17em; margin-top: 1em; margin-bottom: 1em; margin-left: 0; margin-right: 0; font-weight: bold;}
h4 {mso-style-parent:htmlxls2xlsx; display: block; margin-top: 1.33em; margin-bottom: 1.33em; margin-left: 0; margin-right: 0; font-weight: bold; }
h5 {mso-style-parent:htmlxls2xlsx; display: block; font-size: .83em; margin-top: 1.67em; margin-bottom: 1.67em; margin-left: 0; margin-right: 0; font-weight: bold; }
h6 {mso-style-parent:htmlxls2xlsx; display: block; font-size: .67em; margin-top: 2.33em; margin-bottom: 2.33em; margin-left: 0; margin-right: 0; font-weight: bold; }
hr {mso-style-parent:htmlxls2xlsx; border-bottom: 0.5pt solid windowtext; }
i {mso-style-parent:htmlxls2xlsx; font-style: italic;}
p {mso-style-parent:htmlxls2xlsx; display: block; margin-top: 1em; margin-bottom: 1em; margin-left: 0; margin-right: 0;}
pre {mso-style-parent:htmlxls2xlsx; display: block; font-family: monospace; white-space: pre; margin: 1em 0;}
u {mso-style-parent:htmlxls2xlsx; text-decoration: underline;}
s {mso-style-parent:htmlxls2xlsx; text-decoration: line-through;}
small {mso-style-parent:htmlxls2xlsx; font-size: 0.75em;}
span {mso-style-parent:htmlxls2xlsx;}
strike {mso-style-parent:htmlxls2xlsx; text-decoration: line-through;}
strong {mso-style-parent:htmlxls2xlsx; font-weight: bold;}
table {display: table; border-collapse: separate; border-spacing: 2px; border-color: gray; vertical-align: middle; font-family: sans-serif;}
tbody {display: table-row-group; vertical-align: middle; border-color: inherit;}
td {display: table-cell; vertical-align: inherit; white-space: normal; font-size: 1em;}
th {display: table-cell; vertical-align: inherit; font-size: 1em; font-weight: bold; text-align: center; white-space: nowrap;}
thead {display: table-header-group; vertical-align: middle; border-color: inherit;}
tr {display: table-row; vertical-align: inherit; border-color: inherit;}
"""
        self.font_map = {'calibri': 'Calibri', 'arial': 'Arial', 'serif': 'Times New Roman', 
                'monospace': 'Courier New', 'sans-serif': 'Calibri', 'cursive': 'Comic Sans MS'}
        self.number_format_replacements = {'Fixed': r'0.00', 'Number': r'0.00', 'Short Date': r'm\/dd\/yyyy',
                'Long Date': r'dddd\, mmmm d\, yyyy', 'Short Time': r'h\:mm AM/PM', 'Long Time': r'h\:mm\:ss AM/PM',
                'Percent': r'0.00%', 'Scientific': r'0.00E+00'}
        self.pattern_map = {'none': None, 'reverse-diag-stripe': 'darkDown', 'thin-reverse-diag-stripe': 'lightDown',
                'gray-75': 'darkGray', 'gray-25': 'lightGray', 'thick-diag-cross': 'darkTrellis',
                'thin-diag-cross': 'lightTrellis', 'diag-stripe': 'darkUp', 'thin-diag-stripe': 'lightUp',
                'gray-0625': 'gray0625', 'gray-125': 'gray125', 'diag-cross': 'darkGrid',
                'thin-horz-cross': 'lightGrid', 'vert-stripe': 'darkVertical', 'thin-vert-stripe': 'lightVertical',
                'horz-stripe': 'darkHorizontal', 'thin-horz-stripe': 'lightHorizontal', 'gray-50': 'mediumGray',
                'solid': 'solid'}
        self.default_css = {}
        self.add_style_sheet(self.default_styles)
        self.default_css = self.stylemap.copy()
        if TRACE:
            print(f'default_css = {self.default_css}')

    def __str__(self):
        return '\n'.join([f'{style}\n{self.stylemap[style]}' for style in self.stylemap])

    def map_font(self, font):
        """Map the given font to something that Excel probably has"""
        fl = font.lower()
        if fl in self.font_map:
            return self.font_map[fl]
        return font

    @staticmethod
    def _fixup_styles(style):
        if 'mso-number-format' not in style:
            return style
        # Fixup zip-code style which parseStyle changes from 00000 to 0
        style = re.sub(r'mso-number-format\s*:\s*(0[^;]*)', r'mso-number-format:"\1"', style)
        # Fixup mso-number-format style where they forget to add a space after an escape
        # (per css documentation, spaces after an escape sequence are eaten)
        return re.sub(r'\\0022 ', r'\\0022  ', style)

    def add_style_sheet(self, style):
        if not style:
            return

        css = cssutils.parseString(self._fixup_styles(style), validate=False)
        for rule in css:
            if rule.type == rule.STYLE_RULE:
                if ',' in rule.selectorText:    # e1,e2 selects all e1 elements and all e2 element
                    elements = rule.selectorText.split(',')
                    d = self._style_to_dict(rule.style)
                    for element in elements:
                        e = element.strip()
                        self.stylemap[e] = self.update_style(self.stylemap.get(e), d, e)
                elif '>' in rule.selectorText or '[' in rule.selectorText:
                    e = rule.selectorText.strip().replace(' ', '')
                    if '>' in e:
                        tag = e.split('>')[-1]
                    else:
                        tag = None
                    self.stylemap[e] = self.update_style(self.stylemap.get(e), self._style_to_dict(rule.style), tag)
                else:
                    rst = rule.selectorText.strip()
                    rst = rst.replace('tbody', '').replace('thead', '') # We don't send these thru
                    while '  ' in rst:
                        rst = rst.replace('  ', ' ')
                    if rst:     # If it's just 'tbody', now it's empty
                        tag = rst.split()[-1]
                        self.stylemap[rst] = self.update_style(self.stylemap.get(rst), self._style_to_dict(rule.style), tag)

    def _style_to_dict(self, style):
        result = {}
        for item in style:
            if item.name == 'mso-style-parent': # Special Microsoft style inheritance
                if f'.{item.value}' in self.stylemap:
                    parent = self.stylemap[f'.{item.value}']
                    result = self.update_style(parent, result, parent=False)
            else:
                result[item.name] = item.value

        return result

    def _font_size(self, size):
        """Convert the size from a <font size="N"> tag to a point size"""
        result = '12pt'
        try:
            default = 3
            if size[0] == '+':
                size = default + int(size[1:])
            elif size[0] == '-':
                size = default - int(size[1:])
            else:
                size = int(size)
            size = min(max(size, 1), 7)
            result = self.pt_list[size-1]
        except Exception:
            pass
        return result

    def update_style(self, style, new_style, new_tag=None, parent=False):
        """Like normal dict.update() but handle relative size fonts (% and em), inherit, and initial.
        The new_tag specifies the tag name associated with the new_style and is used to process
        values of "initial".  If parent is True (default), then the "style" element is the parent
        style of "new_style", which determines if properties are inherited or just merged.
        """
        if not new_style:
            return style.copy()
        result = {}
        if parent:
            for item, value in style.items():
                if item in self.inherited_properties:
                    result[item] = value
        elif style:
            result = style.copy()
        for item, value in new_style.items():
            if value == 'inherit':
                if item in result:
                    continue
                elif style and item in style:
                    result[item] = style[item]
                    continue
                if self.default_css:
                    value = 'initial'
                elif parent:
                    continue
            if value == 'initial':
                if new_tag in self.default_css and item in self.default_css[new_tag]:
                    value = self.default_css[new_tag][item]
                else:
                    continue
            if item == 'font-size' and (value.endswith('em') or value.endswith('%')) and parent:
                fs = CSSStyle.get_value(value)
                units = CSSStyle.get_units(value)
                if units == '%':
                    fs /= 100
                if 'font-size' in result:
                    pt = CSSStyle.get_pt(result['font-size'])
                else:
                    pt = CSSStyle.DEFAULT_POINT_SIZE
                if pt*fs != 0.0:
                    result['font-size'] = f'{pt*fs}pt'
            else:
                result[item] = value
        return result

    def apply_style(self, tags):
        """Apply the appropriate style to the given element(s).
        The tags is a list of tuples of (tag_name, tag_attrs), given in the proper nested order.

        Handled:
        *
        .class
        #id
        element
        element.class
        element,element     (handled above)
        element element     (direct descendants only)
        element>element
        [attribute]
        [attribute=value]
        
        Not handled:
        .class1.class2
        .class1 .class2
        element+element
        element~element
        """
        master_style = {}
        element_list = []
        for tag_name, attrs in tags:
            style = {}          # Style for this tag
            element_list.append(tag_name)
            if tag_name == 'font':
                color = attrs.get("color")
                face = attrs.get("face")
                size = attrs.get("size")
                styl = ''
                if color:
                    styl += f'color: {color};'
                if face:
                    styl += f'font-family: {face};'
                if size:
                    styl += f'font-size: {self._font_size(size)};'
                css = cssutils.parseStyle(styl)
                style = self.update_style(style, self._style_to_dict(css), tag_name)
            if '*' in self.stylemap:
                style = self.update_style(style, self.stylemap['*'], tag_name)
            if tag_name in self.stylemap:
                style = self.update_style(style, self.stylemap[tag_name], tag_name)
            for i in range(len(element_list)-1):    # table tr td; tr td
                elems = ' '.join(element_list[i:])  # e1 e2 ...
                if elems in self.stylemap:
                    style = self.update_style(style, self.stylemap[elems], tag_name)
                elems = '>'.join(element_list[i:])  # e1>e2 ...
                if elems in self.stylemap:
                    style = self.update_style(style, self.stylemap[elems], tag_name)
            class_ = attrs.get('class')
            if class_ is not None:
                if not isinstance(class_, list):
                    class_ = [class_]
                for c in class_:
                    if f'{tag_name}.{c}' in self.stylemap:
                        style = self.update_style(style, self.stylemap[f'{tag_name}.{c}'], tag_name)
                    if f'.{c}' in self.stylemap:
                        style = self.update_style(style, self.stylemap[f'.{c}'], tag_name)
            id_ = attrs.get('id')
            if id_ is not None and f'#{id_}' in self.stylemap:
                style = self.update_style(style, self.stylemap[f'#{id_}'], tag_name)
            for attr in attrs:
                if f'[{attr}]' in self.stylemap:
                    style = self.update_style(style, self.stylemap[f'[{attr}]'], tag_name)
                if f'[{attr}={attrs[attr]}]' in self.stylemap:
                    style = self.update_style(style, self.stylemap[f'[{attr}={attrs[attr]}]'], tag_name)
                if attr == 'align':
                    style['text-align'] = attrs[attr]
                elif attr == 'valign':
                    style['vertical-align'] = attrs[attr]
            styl = attrs.get('style')
            if styl is not None:
                css = cssutils.parseStyle(self._fixup_styles(styl))
                style = self.update_style(style, self._style_to_dict(css), tag_name)

            master_style = self.update_style(master_style, style, tag_name, parent=True)

        for k, v in list(master_style.items()):     # If any inherit or initial come thru to the end, zap them!
            if v == 'inherit' or v == 'initial':
                del master_style[k]

        return master_style

    def parse_style(self, styl):
        """Parse a style string (e.g. from style="...") to a dict"""
        css = cssutils.parseStyle(self._fixup_styles(styl))
        return self._style_to_dict(css)

    @staticmethod
    def format_style(style):
        """Convert a parsed style dict back to a style string"""
        result = []
        for k, v in style.items():
            result.append(f'{k}:{v}')
        return ';'.join(result)

    @staticmethod
    def to_xlsx_color(color):
        result = Color()        # Basic black
        try:
            if color[0] != '#':
                color = webcolors.name_to_hex(color)
            if len(color) == 4 and color[0] == '#':     # #345 => #334455
                color = color[0] + color[1] + color[1] + color[2] + color[2] + color[3] + color[3]
            result = Color(color[1:])
        except Exception:
            pass
        return result

    @staticmethod
    def get_value(item, default=0.0):
        """For an item like 0.5pt, get the float value = 0.5"""
        m = re.match(r'([+-]?(?:\d+(?:[.]\d*)?)|[.]\d+)', item)
        if m:
            return float(m.group(1))
        return default

    @staticmethod
    def get_units(item, default=None):
        """For an item like 0.5pt, return the units = 'pt'"""
        m = re.match(r'([+-]?(?:\d+(?:[.]\d*)?)|[.]\d+)([^\d]+)$', item)
        if m:
            return m.group(2)
        return default

    @staticmethod
    def get_pt(item, default=0.0, default_units='px', spreadsheet_pt=SPREADSHEET_WIDTH_PX*0.75):
        """For an item like 0.5pt or 2px, get the float value in units of pt"""
        item = str(item)
        result = CSSStyle.get_value(item, default)
        units = CSSStyle.get_units(item, default_units)
        scale = 1.0
        scales = {'px': 0.75, 'em': 11.955168, 'in': 72, 'mm': 2.808, 'cm': 28.08, 'Q': 0.702, 'pc': 12, '%': spreadsheet_pt // 100}
        scale = scales.get(units, scale)
        return result * scale

    @staticmethod
    def get_px(item, default=0.0, spreadsheet_px=SPREADSHEET_WIDTH_PX):
        """For an item like 0.5pt or 2px, get the float value in units of px"""
        result = CSSStyle.get_pt(item, default, spreadsheet_pt=spreadsheet_px*0.75)
        result /= 0.75      # convert pt to px
        return result

    @staticmethod
    def px_to_units(px):
        """Convert pixels to excel column width units (determined empirically)"""
        return px / 7

    @staticmethod
    def units_to_px(units):
        """Convert Excel column width units to pixels"""
        return units * 7


    def style_to_xlsx(self, style, font=None, fill=None, border=None, alignment=None, number_format=None):
        """Convert this style to the appropriate attributes for openpyxl.
        returns a tuple with (font, fill, border, alignment)"""
        if font:
            font = copy.deepcopy(font)
        else:
            font = Font()
        if fill:
            fill = copy.deepcopy(fill)
        else:
            fill = PatternFill()
        if border:
            border = copy.deepcopy(border)
        else:
            border = Border()
        if alignment:
            alignment = copy.deepcopy(alignment)
        else:
            alignment = Alignment()

        if not number_format:
            number_format = 'General'

        def get_side(value):
            val_split = value.split()
            if 'none' in value or 'hidden' in value:
                border_style = None
            elif 'solid' in value:
                border_style = 'medium'
                size = self.get_value(val_split[0], 1.0)
                if size <= 0.25:
                    border_style = 'hair'
                elif size <= 0.5:
                    border_style = 'thin'
                elif size >= 1.5:
                    border_style = 'thick'
            elif 'double' in value:
                border_style = 'double'
            elif 'dotted' in value:
                border_style = 'dotted'
            elif 'dashed' in value:
                border_style = 'dashed'
            color = val_split[-1]
            side = Side(border_style=border_style, color=self.to_xlsx_color(color))
            return side

        for item, value in style.items():
            try:
                if item == 'color':
                    font.color = self.to_xlsx_color(value)
                elif item == 'border':  # 0.5pt solid windowtext
                    side = get_side(value)
                    border = Border(left=side, right=side, top=side, bottom=side)
                elif item == 'border-left-style' or item == 'border-left': # 5px dotted red
                    border.left = get_side(value)
                elif item == 'border-right-style' or item == 'border-right': # 5px dotted red
                    border.right = get_side(value)
                elif item == 'border-top-style' or item == 'border-top': # 5px dotted red
                    border.top = get_side(value)
                elif item == 'border-bottom-style' or item == 'border-bottom': # 5px dotted red
                    border.bottom = get_side(value)
                elif item == 'font-size':
                    sz = self.get_pt(value)
                    if sz:
                        font.sz = sz
                elif item == 'font-weight':     # normal; bold; 400; 700
                    if value == 'bold' or self.get_value(value) > 400:
                        font.b = True
                elif item == 'font-style':      # normal; italic; oblique
                    if value == 'italic' or value == 'oblique':
                        font.i = True
                elif item == 'font-family':     # Arial; Arial, sans-serif; "Comic Sans MS", sans-serif
                    font.name = self.map_font(value.split(',')[0].replace('"', '').replace("'", '').strip())
                    vl = value.lower()
                    if 'sans' in vl or 'calibri' in vl or 'arial' in vl or 'swiss' in vl or \
                      'helvetica' in vl or 'impact' in vl or 'tahoma' in vl or 'veranda' in vl:
                        font.family = 2.0
                    elif 'serif' in vl or 'roman' in vl or 'baskerville' in vl or 'bookman' in vl or \
                      'modern' in vl or 'cambria' in vl or 'century' in vl or 'georgia' in vl:
                        font.family = 1.0
                    elif 'mono' in vl or 'fixed' in vl or 'courier' in vl or 'consol' in vl or \
                      'modem' in vl or 'typewriter' in vl or 'terminal' in vl:
                        font.family = 3.0
                    elif 'script' in vl or 'cursive' in vl:
                        font.family = 4.0
                    elif 'decorative' in vl or 'old english' in vl or 'fraktur' in vl or \
                      'antiqua' in vl:
                        font.family = 5.0
                elif item == 'text-align':      # center
                    alignment.horizontal = value
                elif item == 'text-decoration': # none; underline
                    if 'underline' in value:
                        if font.u != 'double':
                            font.u = 'single'
                    if 'line-through' in value:
                        font.strike = True
                elif item == 'text-underline-style':    # single; double
                    if value == 'double':
                        font.u = 'double'
                elif item == 'vertical-align':  # bottom
                    if value == 'middle':
                        value = 'center'
                    if value in {'distributed', 'justify', 'bottom', 'top', 'center'}:
                        alignment.vertical = value
                elif item == 'white-space': # nowrap
                    if value != 'nowrap':
                        alignment.wrap_text = True
                elif item == 'background':    #this that these those blue
                    value = value.split()[-1]
                    if fill.patternType is None:
                        fill.patternType = 'solid'
                        fill.fgColor = self.to_xlsx_color(value)
                    fill.bgColor = self.to_xlsx_color(value)
                    #fill = PatternFill(patternType='solid', fgColor=self.to_xlsx_color(value), bgColor=self.to_xlsx_color(value))
                elif item == 'background-color':    #f2f2f2
                    #fill = PatternFill(patternType='solid', fgColor=self.to_xlsx_color(value), bgColor=self.to_xlsx_color(value))
                    if fill.patternType is None:
                        fill.patternType = 'solid'
                        fill.fgColor = self.to_xlsx_color(value)
                    fill.bgColor = self.to_xlsx_color(value)
                elif item == 'mso-pattern':     # black gray-50
                    if value != 'auto':
                        value = value.split()
                        fill.fgColor = self.to_xlsx_color(value[0])
                        pattern = 'solid'
                        if len(value) == 2:
                            pattern = value[1]
                        if pattern in self.pattern_map:
                            fill.patternType = self.pattern_map[pattern]
                elif item == 'mso-rotate':
                    if isinstance(value, int) or value.isdigit() or value[0:1] == '-' and value[1:].isdigit(): 
                        rot = int(value)
                        if 1 <= rot <= 180:
                            alignment.textRotation = rot
                        elif -90 <= rot <= -1:
                            alignment.textRotation = 90 - rot
                elif item == 'layout-flow':
                    if 'vertical' in value:
                        alignment.textRotation = 255
                elif item == 'writing-mode':
                    if value == 'tb-rl':
                        alignment.textRotation = 255
                elif item == 'mso-text-control':    # mso-text-control:shrinktofit
                    if value == 'shrinktofit':
                        alignment.shrink_to_fit = True
                elif item == 'mso-char-indent-count':
                    alignment.indent = value
                elif item == 'mso-number-format':   # mso-number-format:"Short Date"
                    if TRACE:
                        print('mso-number-format=', end='')
                        for c in value:
                            print(f'{c}({ord(c):02x})', end='')
                        print('')
                    value = value.strip()
                    if value and value[0] in '"\'' and value[-1] in '"\'':
                        value = value[1:-1]
                    number_format = value.strip()
                    # Examples:
                    # _-* #,##0.00" €"_-;-* #,##0.00" €"_-;_-* -??" €"_-;_-@_-
                    # _-* #,##0.00 [$€-407]_-;-* #,##0.00 [$€-407]_-;_-* "-"?? [$€-407]_-;_-@_-
                    number_format = self.number_format_replacements.get(number_format, number_format). \
                            replace('\\\\', '\\')
                            #replace(r'\\', '').replace(r'\[', '[').replace(r'\]', ']').replace(r'\.', '.'). \
                            #replace(r'\#', '#').replace(r'\/', '/').replace(r'\,', ',')
                    number_format = re.sub(r'\\([.#/,;_"\'\[\]\(\)])', r'\1', number_format)
                    if number_format == r'\@':
                        number_format = '@'
                    if TRACE:
                        print('    number_format=', end='')
                        for c in number_format:
                            print(f'{c}({ord(c):02x})', end='')
                        print('')
            except Exception as e:
                if TRACE:
                    print(f'style_to_xlsx: exception {e} on {item}: {value}')
                    traceback.print_exc()

        return (font, fill, border, alignment, number_format)

    @staticmethod
    def fixup_excel_width(wid):
        # https://foss.heptapod.net/openpyxl/openpyxl/-/issues/293
        if wid >= 1.29:
            result = wid + 0.71
        else:
            result = wid * 1.8
        if result == 0.0:
            return result
        return min(CSSStyle.MAX_CELL_WIDTH_UNITS, max(CSSStyle.px_to_units(CSSStyle.MIN_CELL_WIDTH_PX), result))

    @staticmethod
    def join(base, fn):
        """Join a base URL with a filename or a base local path with a filename"""
        if '://' in base or '://' in fn:   # URL
            result = urljoin(base + '/', fn)
        else:
            result = os.path.join(base, fn)
        if TRACE:
            print(f'join({base}, {fn}) = {result}')
        return result

    @staticmethod
    def read(f, mode='', quiet=False, retries=RETRIES):
        """Read from either a URL or a filename or file-like object.  If mode is 'b', then read in binary.
        If f is a file-like object and you want it read in unicode with the proper encoding, then open it
        using 'rb' mode and don't pass a mode to this method.  If quiet is True, then return None rather 
        than raising an exception on errors"""
        def bytes_to_str(b, filename=None):
            is_html = True
            if b.startswith(b'MIME-Version:'):
                is_html = False
            if filename and filename.endswith('.css'):
                is_html = False
            d = UnicodeDammit(b, is_html=is_html)
            result = d.unicode_markup
            return result

        if isinstance(f, str):
            if '://' in f:  # URL
                for r in range(retries):
                    try:
                        resp = requests.get(f, headers=REQUESTS_HEADER)
                        resp.raise_for_status()
                        if mode == 'b':
                            return resp.content
                        #return resp.text
                        return bytes_to_str(resp.content, f)
                    except Exception:
                        if r == retries-1:
                            if quiet:
                                return None
                            raise
                        sleep(2)
            try:
                with open(f, 'rb') as t:
                    contents = t.read()
                    if mode == 'b':
                        return contents
                    return bytes_to_str(contents, f)
            except Exception:
                if quiet:
                    return None
                raise
        elif isinstance(f, bytes):
            if mode == 'b':
                return f
            return bytes_to_str(f)
        else:
            if 'b' in f.mode and mode != 'b':
                return bytes_to_str(f.read())
            return f.read()

    @staticmethod
    def css_escape_unicode(s):
        def _sub(m):
            ch = ord(m.group(1))
            return f'\\{ch:x} '
        return re.sub(r'([^\x00-\x7F])', _sub, s)

    @staticmethod
    def html_escape_unicode(s):
        def _sub(m):
            ch = ord(m.group(1))
            return f'&#x{ch:x};'
        return re.sub(r'([^\x00-\x7F])', _sub, s)

class HTMLXLS2XLSX:
    """Convert an xls file with html contents into and xlsx file"""
    
    def __init__(self, f, dirname='.'):
        """f is a url, filename, file object, or html string"""
        self.dirname = dirname
        if isinstance(f, str):
            fl = f.lower()
            if '<table' in fl or '<frame' in fl:
                self.text = f
            else:
                self.dirname = os.path.split(f)[0]
                self.text = CSSStyle.read(f)
        else:
            self.text = CSSStyle.read(f)

        self.made_dir = None
        csm = currency_symbols_constants.CURRENCY_SYMBOLS_MAP
        def has_symbol(symbol):
            """Return True iff this symbol is not just a sequence of normal ascii letters, so things like 'KM' will return False but '$b' gives True"""
            m = re.match(r'^[A-Za-z]+$', symbol)
            if m:
                return False
            return True

        currency_symbols_regex = '|'.join([re.escape(symbol) for symbol in csm.values() if has_symbol(symbol)])
        self.RE_INT = re.compile(r'^\s*(?P<sign>[+-]?)(?P<curr_left>\s*(?:' + currency_symbols_regex + r')\s*)?(?P<int>\d+)(?P<curr_right>\s*(?:' + currency_symbols_regex + r'))?\s*$')
        self.RE_FLOAT = re.compile(r'^\s*(?P<sign>[+-]?)(?P<curr_left>\s*(?:' + currency_symbols_regex + r')\s*)?(?P<float>(?:\d+(\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)(?P<curr_right>\s*(?:' + currency_symbols_regex + r'))?\s*$')
        self.RE_YEARFIRST = re.compile(r'(?:yy.*m.*d)|(?:yy.*d.*m)', re.I)
        self.RE_DAYFIRST = re.compile(r'd[d\\/ -]+m', re.I)
        self.EARLIEST_DATE = 1900       # Excel limitation

        if self.text.startswith('MIME-Version:'):
            # Unpack mht file into it's specified tempdir
            msg = email.message_from_string(self.text)
            first_contents = None
            filebase = self.dirname.replace('://', '/')
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                filename = part.get_filename()
                if not filename:
                    # Content-Location: file:///C:/D1A61EE8/width.htm
                    filename = part['Content-Location'].replace('file://', '')
                filename = re.sub(r'/[C-Z]:/', './', filename).replace(':', '')
                if not filename.startswith('.'):        # No root filenames
                    if filename.startswith('/') or filename.startswith('\\'):
                        filename = '.' + filename
                    else:
                        filename = './' + filename
                dirname = os.path.join(filebase, os.path.split(filename)[0])
                try:
                    split_path = re.split(r'[/\\]', dirname)
                    subpaths = [os.path.sep.join(split_path[:i]) for i in range(1, len(split_path)+1)]
                    for subpath in subpaths:
                        if not os.path.exists(subpath):
                            os.makedirs(dirname)
                            if self.made_dir is None:
                                self.made_dir = subpath
                                if TRACE:
                                    print(f'self.made_dir = {self.made_dir}')
                            break
                except FileExistsError:
                    pass
                with open(os.path.join(filebase, filename), 'wb') as fp:
                    cte = part.get('content-transfer-encoding', '').lower()
                    if cte == 'quoted-printable':
                        payload = part.get_payload(decode=False)
                        try:
                            bpayload = payload.encode('ascii')
                        except UnicodeError:
                            # Excel saves "quoted-printable" files that contain non-ascii Unicode
                            # characters, so we need to appropriately encode them in either
                            # css encoding or html encoding style
                            if filename.endswith('css'):
                                payload = CSSStyle.css_escape_unicode(payload)
                            else:
                                payload = CSSStyle.html_escape_unicode(payload)
                            bpayload = payload.encode('ascii')
                        payload = quopri.decodestring(bpayload)
                    else:
                        payload = part.get_payload(decode=True)
                    if not first_contents:
                        first_contents = payload
                        first_filename = filename
                    fp.write(payload)
            if first_contents:
                # The first file is the "workbook" that contains a frameset and frames for each sheet.
                #self.text = str(first_contents, 'utf-8')
                self.text = UnicodeDammit(first_contents, is_html=True).unicode_markup
                self.dirname = os.path.split(os.path.join(filebase, first_filename))[0]

        warnings.filterwarnings("ignore", category=UserWarning, module='bs4')
        warnings.filterwarnings("ignore", category=UserWarning, module='cssutils')
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        warnings.filterwarnings("ignore", category=GuessedAtParserWarning)
        # the newlines are seen as bogus text elements, so get rid of them: (except this messes up <pre>)
        #self.text = ' '.join(self.text.split())
        #self.text = re.sub(r'''<span\s+style=["']mso-spacerun:yes["']>(\s+)</span>''', lambda m: '&nbsp;' * len(m.group(1)), self.text)
        self.text = re.sub(r'>[ \n\t\r\f\v]+<', '><', self.text) 
        #self.text = self.text.replace('> <', '><')
        stl = self.text.lower()
        if '<table' not in stl and '<frame' not in stl:
            if isinstance(f, str):
                raise ValueError(f'No <table> tags found in {f} - maybe this is Excel 5.0/95 format? If so, try using XLS2XLSX instead.')
            else:
                raise ValueError('No <table> tags found - maybe this is Excel 5.0/95 format? If so, try using XLS2XLSX instead.')
        self.url_soup = BeautifulSoup(self.text)
        
    def to_xlsx(self, filename=None, workbook=None, worksheet=None, sheet_name=None):
        """Convert to xlsx using openpyxl.  If filename is not None, then the result
        is written to that file, and the filename is returned, else the workbook is returned.
        If workbook is passed, then the worksheet is written to the given workbook"""
        if workbook:
            wb = workbook
            ws = worksheet
        else:
            wb = Workbook()     # Creates one worksheet
            ws = wb.active

        # Handle a modern "save as htm" file from excel (including converted mht files)
        # <frame src="Styles_files/tabstrip.htm" name="frTabs" marginwidth=0 marginheight=0>
        frames_html = self.url_soup.find_all('frame')
        if frames_html:
            tabstrip = frames_html[-1]
            src = tabstrip['src']
            tabstrip_hx2x = self.__class__(CSSStyle.join(self.dirname, src))
            # <a href="sheet001.htm" target="frSheet"><font face="Arial" color="#000000">Sheet1</font></a>
            a_html = tabstrip_hx2x.url_soup.find_all('a')
            bn = os.path.split(CSSStyle.join(self.dirname, src))[0]
            for a in a_html:
                href = a['href']
                sn = a.get_text().strip()
                fn = CSSStyle.join(bn, href)
                a_hx2x = self.__class__(fn)
                a_hx2x.to_xlsx(workbook=wb, worksheet=ws, sheet_name=sn)
                ws = None
            if self.made_dir:       # If we made the directory from a mht file, then remove it when we're done with it
                try:
                    shutil.rmtree(self.made_dir)
                    if TRACE:
                        print(f'removed {self.made_dir}')
                except Exception as e:
                    if TRACE:
                        print(f'Exception removing {self.made_dir}: {e}')
            if filename:
                wb.save(filename=filename)
                return filename
            return wb

        css_style = CSSStyle()
        font_utils = FontUtils()
    
        styles_html = self.url_soup.find_all(['link', 'style'])
        for s in styles_html:
            if s.name == 'link':
                rel = s.get('rel')
                if not rel or (isinstance(rel, list) and rel[0].lower() != 'stylesheet') or \
                  (isinstance(rel, str) and rel.lower() != 'stylesheet'):
                    continue
                try:
                    fn = s.get('href')
                    cs = CSSStyle.read(CSSStyle.join(self.dirname, fn), quiet=True)
                    if TRACE:
                        print(f'Writing TRACE_{fn}')
                        with open(f'TRACE_{fn}', 'w', encoding='utf-8') as tr:
                            tr.write(cs)
                    css_style.add_style_sheet(cs)
                except Exception:
                    pass
            else:
                css_style.add_style_sheet(str(s.encode_contents(), 'utf-8'))

        # Put elements that are not in a table in a table so we can handle them properly below
        new_row_elems = {'p', 'br', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'hr', 'div', 'pre', 'ol', 'ul', 'dl', 'hr', 'blockquote', 'footer'}

        if TRACE:
            print('Looking for non-table content')
        body = self.url_soup.find('body')
        if not body:
            body = self.url_soup.new_tag('body')
            while self.url_soup.contents:
                elem = self.url_soup.contents[0]
                body.append(elem.extract())
            if 'html' in self.url_soup:
                self.html.append(body)
            else:
                self.url_soup = body
        table = None
        cell = None
        col_widths = {}                 # Measured in units, key is column letter
        col_widths_no_wrap = {}         # Measured in units, key is column letter
        col_max_widths = {}
        while True:
            for elem in body.contents:
                if elem == table:
                    continue
                if isinstance(elem, (CData, Comment, ProcessingInstruction, Declaration, Doctype)):
                    continue
                if isinstance(elem, NavigableString) and not str(elem.string).strip():
                    continue            # Ignore whitespace between elements
                if elem.name == 'table':
                    # If we have an extra table row at the bottom like this one, then use it to set the column
                    # widths, then zap it.
                    """ <![if supportMisalignedColumns]>
                        <tr height=0 style='display:none'>
                        <td width=87 style='width:65pt'></td>
                        <td width=0></td>
                        </tr>
                    <![endif]>"""
                    for e in elem.contents:
                        if isinstance(e, Declaration) and e.string == 'if supportMisalignedColumns':
                            r = e.find_next_sibling()
                            if r.name == 'tr' and 'style' in r.attrs and r.attrs['style'] == 'display:none':
                                cc = 0
                                for c in r.contents:
                                    if c.name != 'td':
                                        continue
                                    cc += 1
                                    if 'width' not in c.attrs:
                                        continue
                                    cl = get_column_letter(cc)
                                    width = CSSStyle.px_to_units(CSSStyle.get_px(c.attrs['width']))
                                    col_widths[cl] = width
                                    col_max_widths[cl] = width
                                    col_widths_no_wrap[cl] = width
                                r.extract()     # Zap it!
                                break
                    table = None            # We need to make a new one if needed
                    continue
                if elem.name in ('link', 'style', 'script'):
                    continue
                if elem.name == 'div' and 'style' in elem.attrs and 'mso-element:comment-list' in elem.attrs['style']:
                    continue        # Ignore the cell comments which are stuck at the end of the file in a div
                if TRACE:
                    print(f'Processing {elem}')
                if not table:
                    if TRACE:
                        print('Creating table')
                    table = self.url_soup.new_tag('table')
                    table.attrs = body.attrs            # Copy bgcolor, style, etc
                    elem.insert_after(table)
                if elem.name in new_row_elems:
                    cell = None
                if not cell:
                    if TRACE:
                        print('Creating new row and cell')
                    row = self.url_soup.new_tag('tr')
                    cell = self.url_soup.new_tag('td')
                    row.append(cell)
                    table.append(row)

                if TRACE:
                    print('Adding element to cell')
                extracted = elem.extract()
                cell.append(extracted)
                break               # Start over if we made a change
            else:
                break

        if TRACE:
            print(f'Styles = \n{css_style}')
    
        def str_size(s, font, alignment, max_width, max_height, keep_newlines=False, fully_merged=False):
            """Get approximate width, height of a string, expressed in pixels."""
            is_bold = font.b
            font_size = font.sz
            if not font_size:
                font_size = 11
            wrap_text = alignment.wrap_text
            rotation = alignment.textRotation
            height = FontUtils.pt_to_px(font_size) * FontUtils.LINE_HEIGHT_FACTOR
            def almost_vertical(rotation):
                # 180 means -90 and 255 means vertical with each letter rotated also
                return 75 <= rotation <= 90 or 165 <= rotation <= 180 or rotation == 255

            if alignment.shrink_to_fit:
                width, _ = font_utils.get_font_size(font, 'a')
            elif wrap_text:
                lines = s.split('\n')
                # Compute the width of the widest line
                mx = 0
                for line in lines:
                    w, h = str_size(line, font, Alignment(wrap_text=False, textRotation=rotation), max_width, max_height)
                    mx = max(mx, w)
                mx = min(mx, max_width)
                my = font_utils.lines_needed(mx, s, font) * height
                # Attempt to optimize the cell size
                width = mx
                height1 = height
                height = my
                if not fully_merged:        # If we don't span the whole width, try to optimize the size
                    if mx/my >= 2:
                        # Speed this up by pre-computing the value based on the area of the rectangle
                        area = width * height
                        # The area of the desired result with a 2:1 aspect ratio is mx * my = (2*my)*my = 2*my**2
                        # so we solve for my.
                        my = math.sqrt(area / 2.0)
                        if my < max_height:
                            mx = 2 * my
                            width = mx
                            my = font_utils.lines_needed(mx, s, font) * height1
                            height = my
                        else:
                            my = height
                    while my < max_height and mx/my >= 2:       # Go for a 2:1 aspect ratio if possible
                        width = mx
                        height = my
                        mx = mx * 0.9
                        my = font_utils.lines_needed(mx, s, font) * height1
                        if TRACE:
                            print(f'str_size: trying {mx, my}')

            elif keep_newlines:
                height = 0
                width = 0
                for line in s.split('\n'):
                    w, h = font_utils.get_font_size(font, line)
                    width = max(width, w)
                    height += h
            else:
                width, height = font_utils.get_font_size(font, s.replace('\n', ''))

            if width > max_width:               # This line will be split
                splits = math.ceil(width / max_width)
                width = max_width
                height *= splits
            if height > max_height:
                height = max_height
            if TRACE:
                print(f'str_size({s}, is_bold={is_bold}, font_size={font_size}, font_name={font.name}, rotation={rotation}, wrap_text={wrap_text}, max_width={max_width}, max_height={max_height}, keep_newlines={keep_newlines}) = {width, height}')
            if almost_vertical(rotation):
                return (height, width)
            return (width, height)
    
        tables_html = self.url_soup.find_all("table")
        if sheet_name is None:
            m = re.search(r'<x:Name>([^<]+?)</x:Name>', self.text)
            if m:
                sheet_name = m.group(1)
            else:
                sheet_name = 'Sheet1'

        if ws:
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        def type_it(value, number_format=None):
            if not value:
                return value
            v = value
            if isinstance(v, float) or isinstance(v, date) or isinstance(v, timedelta):
                return v
            if isinstance(v, int):
                if number_format and ('[h' in number_format or '[m' in number_format or '[s' in number_format):
                    v = str(v)
                else:
                    return v
            if isinstance(v, str):
                v = v.replace(',', '').strip()
                if v[-1:] == '%':
                    try:
                        return round(float(v[:-1]) / 100.0, 15)
                    except Exception:
                        pass
                if number_format and ('[h' in number_format or '[m' in number_format or '[s' in number_format):
                    # timedelta
                    grabit = []
                    def grab_escapes(number_format):
                        nonlocal grabit
                        def sub_grabit(m):
                            i = len(grabit)
                            grabit.append(re.escape(m.group(1)))
                            return f'{{{i}}}'
                        nf = re.sub(r'\\(.)', sub_grabit, number_format)
                        nf = re.sub(r'"([^"]*)"', sub_grabit, nf)
                        return nf

                    def restore_escapes(nf):
                        nonlocal grabit
                        if len(grabit):
                            nf = nf.format(*grabit)       # Put escaped chars back in
                        return nf
                    pattern = grab_escapes(number_format.split(';')[0])
                    pattern = re.sub(r'\[hh?\]', r'(?P<HOURS>\\d+)', pattern)
                    pattern = re.sub(r'\[mm?\]', r'(?P<TOTAL_MINUTES>\\d+)', pattern)
                    pattern = re.sub(r'\[ss?\]', r'(?P<TOTAL_SECONDS>\\d+)', pattern)
                    pattern = re.sub(r'mm?', r'(?P<MINUTES>\\d+)', pattern)
                    pattern = re.sub(r'ss?', r'(?P<SECONDS>\\d+)', pattern)
                    pattern = re.sub(r'[.]00*', r'(?P<FRACTION>\\d+)', pattern)
                    pattern = restore_escapes(pattern)
                    m = re.match(pattern, value)
                    if m:
                        def get(what, default=0):
                            nonlocal m
                            try:
                                return int('0' + m.group(what))
                            except IndexError:
                                return default
                        hours = get('HOURS')
                        total_minutes = get('TOTAL_MINUTES', None)
                        minutes = get('MINUTES')
                        total_seconds = get('TOTAL_SECONDS', None)
                        seconds = get('SECONDS')
                        fraction = float('0.' + str(get('FRACTION')))
                        if TRACE:
                            print(f'type_it({value}, {number_format}) found timedelta: {hours}h{minutes}m{seconds}s {total_minutes}total_m {total_seconds}total_s')
                        if total_seconds is not None:
                            return timedelta(seconds=total_seconds+fraction)
                        if total_minutes is not None:
                            return timedelta(minutes=total_minutes,seconds=seconds+fraction)
                        return timedelta(seconds=(hours*60+minutes)*60+seconds+fraction)
                    elif TRACE:
                        print(f"type_it({value}, {number_format}) timedelta didn't match {pattern.pattern}")

                m = re.match(self.RE_INT, v)
                # Match an integer or float value, possibly having a currency symbol on the left or right (but not both!)
                if m and bool(m.group('curr_left')) + bool(m.group('curr_right')) != 2:
                    return int(m.group('sign') + m.group('int'))
                m = re.match(self.RE_FLOAT, v)
                if m and bool(m.group('curr_left')) + bool(m.group('curr_right')) != 2:
                    return round(float(m.group('sign') + m.group('float')), 15)
            try:
                dayfirst = False
                yearfirst = False
                yearfirst = bool(number_format and re.search(self.RE_YEARFIRST, number_format))
                dayfirst = bool(number_format and re.search(self.RE_DAYFIRST, number_format))
                dt = date_parse(value, dayfirst=dayfirst, yearfirst=yearfirst, default=datetime(1, 1, 1))
                if dt.date() == date(1, 1, 1):
                    return dt.time()
                if dt.year == 1:
                    dt = dt.replace(year, date.today().year)
                if dt.year >= self.EARLIEST_DATE:
                    if dt.time() == tm(0):
                        return dt.date()
                    return dt
            except Exception:
                pass
            if isinstance(value, str):
                return value.strip()
            return value

        def guess_date_format(dt, str_dt):
            """Take a guess at the date format based on a sample in str_dt"""
            nonlocal css_style
            if isinstance(dt, datetime):
                number_format = css_style.number_format_replacements['Short Date'] + ' ' + \
                        css_style.number_format_replacements['Long Time']
                year = dt.year
                month = dt.month
                day = dt.day
                hour = dt.hour
                hour_12 = dt.hour % 12
                if hour_12 == 0:
                    hour_12 = 12
                minute = dt.minute
                second = dt.second
            elif isinstance(dt, date):
                number_format = css_style.number_format_replacements['Short Date']
                year = dt.year
                month = dt.month
                day = dt.day
            elif isinstance(value, tm):
                number_format = css_style.number_format_replacements['Long Time']
                year = 0
                month = 0
                day = 0
                hour = dt.hour
                hour_12 = dt.hour % 12
                if hour_12 == 0:
                    hour_12 = 12
                minute = dt.minute
                second = dt.second
            else:
                return 'General'

            # Tokenize str_dt:
            tokens = []
            nf = ''
            token_seen = set()
            (TOK_HOUR, TOK_HOUR0, TOK_HOUR_12, TOK_HOUR0_12, TOK_MIN, TOK_SEC, TOK_DAY, TOK_DAY0, 
            TOK_MONTH, TOK_MONTH0, TOK_YEAR, TOK_YEAR2, TOK_SHORT_DAY, TOK_LONG_DAY, TOK_SHORT_MONTH, TOK_LONG_MONTH,
            TOK_SEP, TOK_ampm, TOK_AMPM) = range(19)
            token_equiv = [set((TOK_HOUR, TOK_HOUR0, TOK_HOUR_12, TOK_HOUR0_12)),
                    set((TOK_DAY, TOK_DAY0)), set((TOK_MONTH, TOK_MONTH0)),
                    set((TOK_SHORT_DAY, TOK_LONG_DAY)), set((TOK_SHORT_MONTH, TOK_LONG_MONTH)),
                    set((TOK_ampm, TOK_AMPM))]
            token_letters = {TOK_HOUR:'h', TOK_HOUR0:'hh', TOK_HOUR_12:'h', TOK_HOUR0_12:'hh',
                    TOK_MIN:'mm', TOK_SEC:'ss', TOK_DAY:'d', TOK_DAY0:'dd',
                    TOK_MONTH:'m', TOK_MONTH0:'mm', TOK_YEAR:'yyyy', TOK_YEAR2:'yy',
                    TOK_SHORT_DAY:'ddd', TOK_LONG_DAY:'dddd', TOK_SHORT_MONTH:'mmm',
                    TOK_LONG_MONTH:'mmmm', TOK_ampm:'a/p\\m', TOK_AMPM:'AM/PM'}
            str_dtl = str_dt.lower()
            has_am_pm = 'am' in str_dtl or 'pm' in str_dtl
            toks = re.split(r'(\W+|(?:(?<![A-Za-z_])(?:am\b|pm\b|AM\b|PM\b))|(?:(?<=[0-9])[A-Za-z_]+)|(?:(?<=[A-Za-z_])[0-9]+))', str_dt)
            matches = 0
            for i, tok in enumerate(toks):
                if not tok:
                    continue
                # Look ahead to mark this as a time if we see a ':' or am/pm or space then am/pm
                look_ahead = ''
                for j in range(i+1, len(toks)):
                    look_ahead += toks[j]
                    if len(look_ahead) >= 2:
                        break
                look_ahead_time = (look_ahead[0:1] == ':' or \
                  look_ahead[0:2].lower() in ('am', 'pm') or \
                  look_ahead[0:3].lower() in (' am', ' pm'))
                if tok.isdigit():
                    t_type = TOK_SEP
                    if tok == str(year):
                        t_type = TOK_YEAR
                    elif TOK_MONTH not in token_seen and tok == str(month) and not look_ahead_time:
                        t_type = TOK_MONTH
                    elif TOK_MONTH0 not in token_seen and month < 10 and tok == '0' + str(month) and \
                      not look_ahead_time:
                        t_type = TOK_MONTH0
                    elif TOK_DAY not in token_seen and tok == str(day) and not look_ahead_time:
                        t_type = TOK_DAY
                    elif TOK_DAY0 not in token_seen and day < 10 and tok == '0' + str(day) and not look_ahead_time:
                        t_type = TOK_DAY0
                    elif tok == str(year%100) and not look_ahead_time:
                        t_type = TOK_YEAR2
                    elif TOK_HOUR not in token_seen and tok == str(hour) and not has_am_pm:
                        t_type = TOK_HOUR
                    elif TOK_HOUR0 not in token_seen and hour < 10 and tok == '0' + str(hour) and not has_am_pm:
                        t_type = TOK_HOUR0
                    elif TOK_HOUR_12 not in token_seen and tok == str(hour_12) and has_am_pm:
                        t_type = TOK_HOUR_12
                    elif TOK_HOUR0_12 not in token_seen and hour_12 < 10 and tok == '0' + str(hour_12) and has_am_pm:
                        t_type = TOK_HOUR0_12
                    elif TOK_MIN not in token_seen and tok == f'{minute:02d}':
                        t_type = TOK_MIN
                    elif TOK_SEC not in token_seen and tok == f'{second:02d}':
                        t_type = TOK_SEC
                elif tok in calendar.day_abbr:
                    t_type = TOK_SHORT_DAY
                elif tok in calendar.day_name:
                    t_type = TOK_LONG_DAY
                elif tok in calendar.month_abbr:
                    t_type = TOK_SHORT_MONTH
                elif tok in calendar.month_name:
                    t_type = TOK_LONG_MONTH
                elif tok in ('am', 'pm'):
                    t_type = TOK_ampm
                elif tok in ('AM', 'PM'):
                    t_type = TOK_AMPM
                else:
                    t_type = TOK_SEP
                    #tok = re.sub(r'([hHmMsSdDyYaApP])', r'\\\1', tok)  # Escape pattern letters
                    tok = re.sub(r'([A-Za-z0-9])', r'\\\1', tok)  # Escape all letters and numbers
                t_val = token_letters.get(t_type, tok)
                if t_val is not tok:
                    matches += 1        # Gotta see 2 good ones to call "success"
                tokens.append((t_type, t_val))
                token_seen.add(t_type)
                for te in token_equiv:
                    if t_type in te:
                        for t in te:
                            token_seen.add(t)
                        break
            for t_type, t_val in tokens:
                nf += t_val
            if matches >= 2:
                number_format = nf

            if TRACE:
                print(f'guess_date_format({dt}, {str_dt}) = {number_format} (toks={toks})')
            return number_format
        
        def get_heights(elem, style):
            """Return the height, min_height, max_height of a given element with the given style"""
            height = None       # height could be 0 for hidden rows
            min_height = None
            max_height = CSSStyle.MAX_CELL_HEIGHT_PT
            if 'min-height' in style:
                min_height = CSSStyle.get_pt(style['min-height'])
            if 'max-height' in style:
                max_height = CSSStyle.get_pt(style['max-height'])
            if 'height' in style:
                height = CSSStyle.get_pt(style['height'])
            elif 'min-height' in style:
                height = CSSStyle.get_pt(style['min-height'])
            elif 'height' in row.attrs:
                height = CSSStyle.get_pt(row['height'], spreadsheet_pt=CSSStyle.SPREADSHEET_HEIGHT_PX*0.75)
            return (height, min_height, max_height)

        global_n_cols = 0
        for table in tables_html:
            parent = table.parent
            while parent:
                if parent.name == 'td':
                    break
                parent = parent.parent
            if parent and parent.name == 'td':
                continue
            for row in table.find_all("tr"):
                if row.parent != table and row.parent.parent != table:     # Don't look at nested rows (2 checks because of thead/tbody)
                    continue
                col_tags = row.find_all(["td", "th"])
                if len(col_tags) > 0:
                    cols = 0
                    for col in col_tags:
                        if col.parent != row:
                            continue
                        colspan = col.get("colspan")
                        if colspan is None:
                            cs = 1
                        else:
                            cs = int(colspan)
                        cols += cs
                    if cols > global_n_cols:
                        global_n_cols = cols
                        if TRACE:
                            print(f'Row {row} set global_n_cols to {global_n_cols}')

        if TRACE:
            print(f'global_n_cols = {global_n_cols}')

        def style_without(style, what='mso-ignore'):
            result = {}
            for k, v in style.items():
                if k == what:
                    continue
                result[k] = v
            return CSSStyle.format_style(result)

        # Handle mso-xlrowspan:N, which means to pretend there are more rows than specified by <tr>
        #  <tr height=3D40 style=3D'height:30.0pt;mso-xlrowspan:2'>
        # We insert new rows and split the height up amongst them.

        xlrowspans = self.url_soup.find_all('tr', style=lambda s: s and 'mso-xlrowspan' in s)
        for elem in xlrowspans:
            style = css_style.parse_style(elem.get('style'))
            rowspan = int(style['mso-xlrowspan'])
            height = None
            if 'height' in style:
                height = CSSStyle.get_pt(style['height'])
            else:
                height = elem.get('height')
            if height is not None:
                height = int(height) // rowspan
                elem['height'] = height
                style['height'] = f'{height}pt'
            elem['style'] = style_without(style, what='mso-xlrowspan')
            for r in range(rowspan-1):
                new_tr = copy.copy(elem)    # This actually calls a custom __copy__ in bs4
                elem.insert_after(new_tr)

        # Handle mso-ignore:colspan; mso-ignore:style; mso-ignore:color; mso-ignore:colspan-rowspan
        # by re-writing the html.  Examples:
        # <td colspan=7 rowspan=15 height=300 width=560 style='mso-ignore:colspan-rowspan;height:225.0pt;width:420pt'>CHART</td>
        # <td height=20 class=xl79 colspan=2 style='height:15.0pt;mso-ignore:colspan'>Shrink to fit text</td>
        # <td height=20 colspan=2 style='height:15.0pt;mso-ignore:colspan'></td>
        # <font color="#0000FF" style='mso-ignore:color'>47712</font>

        def element_index(elem):
            parent = elem.parent
            for i, e in enumerate(parent.contents):
                if e == elem:
                    return i
            return None

        ignores = self.url_soup.find_all(style=lambda s: s and 'mso-ignore' in s)
        for elem in ignores:
            style = css_style.parse_style(elem.get('style'))
            what = style.get('mso-ignore')
            if what == 'colspan' or what == 'colspan-rowspan':
                colspan = int(elem.get("colspan"))
                width = None
                if 'width' in style:
                    width = CSSStyle.get_px(style['width'])
                else:
                    width = elem.get("width")
                if width is not None:
                    width = int(width) // colspan
                    elem['width'] = width
                    style['width'] = f'{width}px'
                # Leave it in so we can use it later not to set the column width: elem['style'] = style_without(style)
                if what == 'colspan-rowspan':
                    elem['style'] = elem['style'] + ';mso-ignore:rowspan'
                del elem['colspan']
                for c in range(colspan-1):
                    new_td = self.url_soup.new_tag(elem.name)
                    new_td.attrs = elem.attrs.copy()
                    elem.insert_after(new_td)
            else:                   # color or style or whatever
                try:
                    elem['style'] = style_without(style)
                    del elem[what]
                except Exception:
                    pass
        ignores = self.url_soup.find_all(['td', 'th'], style=lambda s: s and 'mso-ignore:rowspan' in s)
        for elem in ignores:
            style = css_style.parse_style(elem.get('style'))
            rowspan = int(elem.get('rowspan'))
            height = None
            if 'height' in style:
                height = CSSStyle.get_pt(style['height'])
            else:
                height = elem.get('height')
            if height is not None:
                height = int(height) // rowspan
                elem['height'] = height
                style['height'] = f'{height}pt'
            elem['style'] = style_without(style)
            del elem['rowspan']
            ndx = element_index(elem)
            row = elem.parent
            for r in range(rowspan-1):
                new_td = self.url_soup.new_tag(elem.name)
                new_td.attrs = elem.attrs.copy()
                row = row.find_next_sibling('tr')
                if ndx == 0:
                    e = row.contents[0]
                    e.insert_before(new_td)
                else:
                    e = row.contents[ndx-1]
                    e.insert_after(new_td)

        # Code based on the work of John Ricco (johnricco226@gmail.com), Apr 4, 2017:
        # https://johnricco.github.io/2017/04/04/python-html/

        # Parse each table
        row_counter = 0
        row_heights = {}                # Measured in pt
        row_heights_per_column_no_wrap = {}
        row_heights_per_column = {}
        has_merged_cell = False
        fully_merged_rows = set()
        extra_images = []   # If we have more than 1 image per cell, we put them here until the end
        has_image = set()   # Cells (like A1) that have images
    
        for table in tables_html:
            parent = table.parent
            while parent:
                if parent.name == 'td':  # Nested tables are handled (implicitly) below
                    break
                parent = parent.parent
            if parent and parent.name == 'td':
                continue
            pf = None
            bgcolor = table.get('bgcolor')
            if bgcolor:
                color = CSSStyle.to_xlsx_color(bgcolor)
                pf = PatternFill(patternType='solid', fgColor=color, bgColor=color)
            bor = None
            b = type_it(table.get('border'))
            if b:
                if b == 1:
                    border_style = 'thin'
                else:
                    border_style = 'medium'
                side = Side(border_style=border_style)
                bor = Border(left=side, right=side, top=side, bottom=side)
            
            # Create list to store rowspan values 
            skip_index = [0 for i in range(0, global_n_cols)]
            if TRACE:
                print(f'skip_index = {skip_index}')
                    
            # Start by iterating over each row in this table...
            for row in table.find_all("tr"):
                if row.parent != table and row.parent.parent != table:     # Don't look at nested rows (2 checks because of thead/tbody)
                    continue
                rw = row_counter+1
                # Skip row if it's blank
                columns = row.find_all(["td", "th"])
                lc = len(columns)
                if lc == 0:
                    continue
                else:
                    row_heights_per_column[rw] = {}
                    row_heights_per_column_no_wrap[rw] = {}
                    tuples = [('table', table.attrs), ('tr', row.attrs)]
                    style = css_style.apply_style(tuples)


                    height, min_height, max_height = get_heights(row, style)

                    if height is not None:
                        if rw in row_heights:
                            row_heights[rw] = max(row_heights[rw], height)
                        else:
                            row_heights[rw] = height
                        if min_height and row_heights[rw] < min_height:
                            row_heights[rw] = min_height
                        if max_height and row_heights[rw] > max_height:
                            row_heights[rw] = max_height

                    # Get all cells containing data in this row
                    col_dim = []
                    row_dim = []
                    col_dim_counter = -1
                    row_dim_counter = -1
                    col_counter = -1
                    this_skip_index = copy.deepcopy(skip_index)
                    if len(columns) < global_n_cols:
                        # Excel is one big table, so if we have multple tables in the html,
                        # we must merge them into one, and if the rows have a different #
                        # of columns, we set the colspan on the last column to make up the
                        # difference.
                        col_count = 0
                        for si in skip_index:
                            if si > 0:
                                col_count += 1
                        for col in columns:
                            if col.parent != row:
                                continue
                            colspan = col.get("colspan")
                            if colspan is None:
                                cs = 1
                            else:
                                cs = int(colspan)
                            col_count += cs
                        if col_count != global_n_cols:
                            colspan = columns[-1].get("colspan")
                            if colspan is None:
                                cs = 1
                            else:
                                cs = int(colspan)
                            if TRACE:
                                print(f'setting colspan for last col on row {rw} to {cs+global_n_cols-col_count}')
                            columns[-1]['colspan'] = cs + (global_n_cols - col_count)

                    images = []
                    for col in columns:
                        if col.parent != row:
                            continue
                        tuples = [('table', table.attrs), ('tr', row.attrs), (col.name, col.attrs)]
                        alg = None
                        if 'align' in col.attrs:
                            alg = Alignment(horizontal=col.attrs['align'])
                        if 'valign' in col.attrs:
                            if value == 'middle':
                                value = 'center'
                            if value in {'distributed', 'justify', 'bottom', 'top', 'center'}:
                                if alg:
                                    alg.vertical = value
                                else:
                                    alg = Alignment(vertical=value)
                        # If this cell contains one formatting tag, then apply that tag to the entire contents of
                        # the cell, since openpyxl doesn't support rich text.  If that tag also contains one
                        # formatting tag, then keep going, so <td><font size="5"><span style="font-family: monospace">...</span></font></td> will apply
                        # both styles to the entire cell.
                        c_col = col
                        while c_col.contents and len(c_col.contents) == 1 and c_col.contents[0].name in \
                          ('big', 'small', 'center', 'div', 'span', 'a', 'b', 'i', 'u', 'em', 's', 'strike', 
                           'code', 'pre', 'strong', 'font', 'p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'hr'):
                            cont = c_col.contents[0]
                            tuples.append((cont.name, cont.attrs))
                            c_col = cont
                        style = css_style.apply_style(tuples)
                        font, fill, border, alignment, number_format = css_style.style_to_xlsx(style, fill=pf, border=bor, alignment=alg)
                        font_size = font.sz
                        if not font_size:
                            font_size = 11
                        padding_left = (font_size / 10)        # in px
                        padding_right = padding_left
                        padding_top = padding_left * 0.75      # in pt = 1px
                        padding_bottom = padding_top
                        if style:
                            padding_left = CSSStyle.get_px(style.get('padding-left', '1px'))
                            padding_right = CSSStyle.get_px(style.get('padding-right', '1px'))
                            padding_top = CSSStyle.get_pt(style.get('padding-top', '1pt'))
                            padding_bottom = CSSStyle.get_pt(style.get('padding-bottom', '2pt'))
                            if 'padding' in style:      # 4 in 1
                                paddings = style['padding'].split()
                                if len(paddings) >= 4:
                                    padding_top = CSSStyle.get_pt(paddings[0])
                                    padding_right = CSSStyle.get_px(paddings[1])
                                    padding_bottom = CSSStyle.get_pt(paddings[2])
                                    padding_left = CSSStyle.get_px(paddings[2])
                                elif len(paddings) == 3:
                                    padding_top = CSSStyle.get_pt(paddings[0])
                                    padding_right = CSSStyle.get_px(paddings[1])
                                    padding_left = padding_right
                                    padding_bottom = CSSStyle.get_pt(paddings[2])
                                elif len(paddings) == 2:
                                    padding_top = CSSStyle.get_pt(paddings[0])
                                    padding_bottom = padding_top
                                    padding_right = CSSStyle.get_px(paddings[1])
                                    padding_left = padding_right
                                elif len(paddings) == 1:
                                    padding_top = CSSStyle.get_pt(paddings[0])
                                    padding_bottom = padding_top
                                    padding_right = CSSStyle.get_px(paddings[0])
                                    padding_left = padding_top

                        padding_left = CSSStyle.px_to_units(padding_left)
                        padding_right = CSSStyle.px_to_units(padding_right)
                        font_name = font_utils.get_real_font_name(font.name)
                        if font_name:
                            font.name = font_name

                        # Determine cell dimensions
                        colspan = col.get("colspan")
                        if colspan is None:
                            col_dim.append(1)
                        else:
                            col_dim.append(int(colspan))
                        col_dim_counter += 1
                            
                        rowspan = col.get("rowspan")
                        if rowspan is None:
                            row_dim.append(1)
                        else:
                            row_dim.append(int(rowspan))
                        row_dim_counter += 1
                            
                        # Adjust column counter
                        if col_counter == -1:
                            col_counter = 0  
                        else:
                            col_counter = col_counter + col_dim[col_dim_counter - 1]
                            
                        while skip_index[col_counter] > 0:
                            col_counter += 1

                        cc = col_counter+1
                        cl = get_column_letter(cc)

                        if TRACE and style:     # pragma nocover
                            print(f'style={style} for {cl}{rw} with {col.get_text()}')
                            print(f'font={font}, fill={fill}, border={border} ({border.left}, {border.top}, {border.right}, {border.bottom}), alignment={alignment}, number_format={number_format}')

                        # Get cell contents  
                        #cell_data = col.get_text()     # It's not quite this easy!!
                        white_space = style.get('white-space', 'nowrap' if col.name=='th' else 'normal')
                        cell_data = ''
                        hyperlink = None
                        comment_author = None
                        comment_contents = None
                        contents = list(col.contents)
                        i = -1
                        while True:
                            i += 1
                            if i >= len(contents):
                                break
                            cont = contents[i]
                            if isinstance(cont, (CData, Comment, ProcessingInstruction, Declaration, Doctype)):
                                continue
                            elif isinstance(cont, NavigableString):
                                if white_space == 'pre':
                                    cell_data += str(cont.string)
                                else:
                                    cont_s = str(cont.string)
                                    cont_s = cont_s.replace('\r', '')
                                    cont_s = re.sub(r'\n[ \n\r\t\f\v]*', ' ', cont_s)    # Don't match non-breaking space with \s
                                    if cell_data.endswith('\n') or cell_data.endswith(' '):
                                        cont_s = cont_s.lstrip()                    # Doesn't remove non-breaking spaces
                                    cell_data += cont_s.replace('\xA0', ' ')        # Non-breaking space to space
                            else:
                                css = css_style.apply_style([(cont.name, cont.attrs)])
                                if css.get('display') == 'none':
                                    continue
                                elif css.get('visibility') == 'hidden':
                                    text = cont.get_text()
                                    cell_data += ' ' * len(text)
                                    continue
                                if cont.name == 'br':
                                    cell_data += '\n'
                                elif cont.name == 'p':
                                    if cell_data:
                                        cell_data += '\n'
                                elif cont.name == 'a':
                                    if isinstance(hyperlink, str):
                                        hyperlink = False       # We can only handle 1 per cell
                                    elif hyperlink is None:
                                        hyperlink = cont.get('href')
                                        if hyperlink and hyperlink[0] == '#':
                                            hyperlink = None   # Ignore relative references
                                elif cont.name == 'span' and 'onmouseover' in cont.attrs:   # msoCommentShow('_com_1','_anchor_1')
                                    m = re.match(r"msoCommentShow\('([^']+)'", cont.attrs['onmouseover'])
                                    try:
                                        id_ = m.group(1)
                                        comment_div = self.url_soup.find(id=id_).div.div
                                        comment_author = comment_div.contents[0].get_text()
                                        comment_contents = comment_div.contents[1].get_text()
                                    except Exception:
                                        pass
                                elif cont.name == 'img':
                                    src = cont.get('src')
                                    if src:
                                        try:
                                            content = CSSStyle.read(CSSStyle.join(self.dirname, src), mode='b', retries=2)
                                            image = Image(io.BytesIO(content))
                                            images.append(image)
                                            # If this img is in a nested table (e.g. from a chart), then don't update the heights and widths
                                            parent = cont.parent
                                            while parent:
                                                if parent.name == 'table':
                                                    break
                                                elif parent.name == 'span' and \
                                                  'position:absolute' in parent.attrs.get('style', ''):
                                                    # If this is positioned, then it doesn't change the cell size
                                                    parent = None
                                                    break
                                                parent = parent.parent
                                            if parent == table:
                                                height = cont.get('height')
                                                if height:
                                                    height = CSSStyle.get_pt(height) + padding_top + padding_bottom
                                                    if rw in row_heights:
                                                        row_heights[rw] = max(row_heights[rw], height)
                                                    else:
                                                        row_heights[rw] = height
                                                width = cont.get('width')
                                                if width:
                                                    width = CSSStyle.px_to_units(CSSStyle.get_px(width))
                                                    col_widths[cl] = max(col_widths.get(cl, 0), width+padding_left+padding_right)
                                                    col_widths_no_wrap[cl] = max(col_widths_no_wrap.get(cl, 0), width+padding_left+padding_right)
                                        except Exception as e:
                                            if TRACE:
                                                print(f'Exception getting image for {src}: {e}')

                                # Depth first traversal: insert all sub-nodes in the list after this one
                                contents = contents[0:i+1] + list(cont.contents) + contents[i+1:]
                        
                        # Insert data into cell
                        if TRACE:
                            print(f'ws.cell({rw}, {cc}).value = {type_it(cell_data, number_format)}')
                        try:
                            value = cell_data
                            if number_format != '@' and number_format != 'Text':
                                value = type_it(cell_data, number_format)
                            ws.cell(rw, cc).value = value
                            if number_format == 'General' and isinstance(value, (date, tm)):
                                number_format = guess_date_format(value, cell_data)
                            elif isinstance(cell_data, str):
                                if '\n' in cell_data and not alignment.wrap_text:
                                    alignment = copy.deepcopy(alignment)
                                    alignment.wrap_text = True
                                    white_space = 'normal'
                                if cell_data[-1:] == '%' and number_format == 'General':
                                    number_format = numbers.FORMAT_PERCENTAGE
                            ws.cell(rw, cc).font = font
                            ws.cell(rw, cc).fill = fill
                            ws.cell(rw, cc).border = border
                            ws.cell(rw, cc).alignment = alignment
                            ws.cell(rw, cc).number_format = number_format
                            if hyperlink:
                                ws.cell(rw, cc).hyperlink = hyperlink
                                #ws.cell(rw, cc).style = 'Hyperlink'  # Added a default style instead
                            if images:
                                # Even though the cell may have multiple images in it, we only add one
                                # to each cell, and the rest go in subsequent cells, handled later
                                image = images.pop(0)       # Pop off the first one
                                image.anchor = f'{cl}{rw}'
                                has_image.add(image.anchor)
                                ws.add_image(image)
                                width = image.width
                                for image in images:
                                    image_data = (rw, cc, width, image)
                                    extra_images.append(image_data) # Save for later
                                    width += image.width
                                images = []
                            if comment_contents and comment_author:
                                ws.cell(rw, cc).comment = OpenpyxlComment(comment_author + comment_contents, 
                                                                          comment_author[:-1])   # Remove the ':' from author
                        except AttributeError as e:     # pragma nocover
                            if TRACE:           # Shouldn't happen anymore, but don't crash if it does
                                print(f'AttributeError {e} (ignored)')
                        cd = col_dim[-1]
                        rd = row_dim[-1]
                        # mso-ignore now handled above
                        #if style and 'colspan' in style.get('mso-ignore', ''):
                            #cd = 1
                        #if style and 'rowspan' in style.get('mso-ignore', ''):
                            #rd = 1
                        rx = rw
                        if cd != 1 or rd != 1:
                            has_merged_cell = True
                            if TRACE:
                                print(f'ws.merge_cells(start_row={rw}, start_column={cc}, end_row={row_counter+rd}, end_column={col_counter+cd})')
                            end_row = row_counter+rd
                            ws.merge_cells(start_row=rw,
                                    start_column=cc, 
                                    end_row=end_row,
                                    end_column=col_counter+cd)
                            if col_counter+cd >= global_n_cols:
                                if cc == 1:
                                    fully_merged_rows.add(rw)
                                    row_counter += rd-1
                            if cd != 1:
                                cl = f'{cl}:{get_column_letter(col_counter+cd)}'
                            if rd != 1:
                                rx = f'{rw}:{end_row}'
                                height, min_height, max_height = get_heights(col, style)
                                if height is not None:
                                    row_heights[rx] = height
                                row_heights_per_column[rx] = {}
                                row_heights_per_column_no_wrap[rx] = {}

                        width = None
                        min_width_px = None
                        max_width_px = CSSStyle.units_to_px(CSSStyle.MAX_CELL_WIDTH_UNITS)
                        if 'min-width' in style:
                            min_width_px = CSSStyle.get_px(style['min-width'])
                        if 'max-width' in style:
                            max_width_px = CSSStyle.get_px(style['max-width'])
                        if 'width' in style:
                            width = CSSStyle.px_to_units(CSSStyle.get_px(style['width']))
                        elif 'min-width' in style:
                            width = CSSStyle.px_to_units(CSSStyle.get_px(style['min-width']))
                        elif 'width' in col.attrs:
                            w = col['width']
                            width = CSSStyle.px_to_units(CSSStyle.get_px(w))
                        # mso-ignore:colspan or mso-ignore:colspan-rowspan in a cell with contents
                        # means it's OK for the contents to run over into the next column
                        ignore_colspan = 'mso-ignore' in style and 'colspan' in style['mso-ignore']
                        if width is not None and not ignore_colspan:
                            col_widths[cl] = max(col_widths.get(cl, 0), width)
                            col_widths_no_wrap[cl] = max(col_widths_no_wrap.get(cl, 0), width)
                        if col_widths.get(cl) != 0.0 and not ignore_colspan:    # 0 means it was hidden, and that's usually set on the first <td> for that column
                            # Grab the heights for later so we can decide to use either the wrapped height or the non-wrapped
                            height = row_heights.get(rx, -1)
                            if height != 0:     # 0 means this row is hidden, so just skip it
                                max_height_px = FontUtils.pt_to_px(max_height)
                                wi, he = str_size(str(cell_data).strip(), font, alignment, max_width_px, max_height_px, fully_merged=rw in fully_merged_rows)
                                row_heights_per_column[rx][cl] = max(height, FontUtils.px_to_pt(he)+padding_top+padding_bottom)
                                wiu = CSSStyle.px_to_units(wi)
                                col_widths[cl] = max(col_widths.get(cl, 0), wiu+padding_left+padding_right)
                                if white_space == 'nowrap':
                                    col_widths_no_wrap[cl] = col_widths[cl]
                                    row_heights_per_column_no_wrap[rx][cl] =  row_heights_per_column[rx].get(cl, 0)
                                else:
                                    al = copy.deepcopy(alignment)
                                    al.wrap_text = False
                                    wi, he = str_size(str(cell_data).strip(), font, al, max_width_px, max_height_px, keep_newlines=True)
                                    row_heights_per_column_no_wrap[rx][cl] = max(height, FontUtils.px_to_pt(he)+padding_top+padding_bottom)
                                    wiu = CSSStyle.px_to_units(wi)
                                    col_widths_no_wrap[cl] = max(col_widths_no_wrap.get(cl, 0), wiu+padding_left+padding_right)
                        if max_width_px and cl in col_widths:
                            max_width = CSSStyle.px_to_units(max_width_px)
                            col_widths[cl] = min(col_widths.get(cl), max_width)
                            col_widths_no_wrap[cl] = min(col_widths_no_wrap.get(cl), max_width)
                            if cl not in col_max_widths:
                                col_max_widths[cl] = max_width

                        # Record column skipping index
                        if row_dim[row_dim_counter] > 1:
                            rd = row_dim[row_dim_counter]
                            for i in range(cd):
                                if TRACE:
                                    print(f'this_skip_index[{col_counter+i}] = row_dim[{row_dim_counter}] = {rd}')
                                this_skip_index[col_counter+i] = row_dim[row_dim_counter]
                
                # Adjust row counter 
                row_counter += 1
                
                # Adjust column skipping index
                skip_index = [i - 1 if i > 0 else i for i in this_skip_index]
                if TRACE:
                    print(f'skip_index = {skip_index}')

        def handle_merged_column_widths(col_widths):
            for col, wid in list(col_widths.items()):
                cs = col.split(':')
                if len(cs) != 2:
                    continue
                c1 = column_index_from_string(cs[0])
                c2 = column_index_from_string(cs[1])
                # We have 3 cases for the columns covered by this range:
                # (1) None of the column widths are specified.  In this case divide up this width evenly amongst the columns.
                # (2) Some of the column widths are specified.  In this case, divide up the remaining width evenly amongst the remaining columns.
                # (3) All of the column widths are specified.  In this case, compute the ratio between the total widths of
                #     the columns covered, and use that to bump up (not down) each width appropriately.
                no_width_cols = set()
                total_width = 0
                for c in range(c1, c2+1):
                    l = get_column_letter(c)
                    if l in col_widths:
                        total_width += col_widths[l]
                    else:
                        no_width_cols.add(c)
                remaining_width = wid - total_width
                if remaining_width > 0:
                    ncols = len(no_width_cols)
                    if ncols:           # Case 1 or Case 2
                        width_per = remaining_width / ncols
                        for c in no_width_cols:
                            col_widths[get_column_letter(c)] = width_per
                    elif total_width != 0:       # Case 3
                        ratio = wid / total_width
                        for c in range(c1, c2+1):
                            col_widths[get_column_letter(c)] *= ratio
                del col_widths[col]

        def handle_merged_row_heights(row_heights):
            for row, hei in list(row_heights.items()):
                if isinstance(hei, dict):       # This is a row_heights_per_column
                    # If we have a height specified for a set of merged columns, then just apply it to each column separately
                    for cc, h in list(hei.items()):
                        cs = cc.split(':')
                        if len(cs) == 2:
                            c1 = column_index_from_string(cs[0])
                            c2 = column_index_from_string(cs[1])
                            for c in range(c1, c2+1):
                                l = get_column_letter(c)
                                if l in hei:
                                    hei[l] = max(hei[l], h)
                                else:
                                    hei[l] = h
                            del hei[cc]

                    # At this point, each hei item is for a single column
                    if isinstance(row, str):
                        for cc, h in hei.items():
                            rh = {}
                            for r, d in row_heights.items():
                                if cc in d:
                                    rh[r] = d[cc]
                            handle_merged_row_heights(rh)
                            for r in rh:
                                if r not in row_heights:
                                    row_heights[r] = {}
                                row_heights[r][cc] = rh[r]
                        del row_heights[row]
                    continue

                if not isinstance(row, str):
                    continue
                rs = row.split(':')
                if len(rs) != 2:
                    continue
                r1 = int(rs[0])
                r2 = int(rs[1])

                # We have 3 cases for the rows covered by this range:
                # (1) None of the row heights are specified.  In this case divide up this height evenly amongst the rows.
                # (2) Some of the row heights are specified.  In this case, divide up the remaining height evenly amongst the remaining rows.
                # (3) All of the row heights are specified.  In this case, compute the ratio between the total height of
                #     the rows covered, and use that to bump up (not down) each height appropriately.
                no_height_rows = set()
                total_height = 0
                for r in range(r1, r2+1):
                    if r in row_heights:
                        total_height += row_heights[r]
                    else:
                        no_height_rows.add(r)
                remaining_height = hei - total_height
                if remaining_height > 0:
                    nrows = len(no_height_rows)
                    if nrows:           # Case 1 or Case 2
                        height_per = remaining_height / nrows
                        for r in no_height_rows:
                            row_heights[r] = height_per
                    elif total_height != 0:       # Case 3
                        ratio = hei / total_height
                        for r in range(r1, r2+1):
                            row_heights[r] *= ratio
                del row_heights[row]

        if has_merged_cell:
            if TRACE:
                print("Values before handling merged cells:")
                print(f'col_widths                     = {col_widths}')
                print(f'col_widths_no_wrap             = {col_widths_no_wrap}')
                print(f'row_heights                    = {row_heights}')
                print(f'row_heights_per_column         = {row_heights_per_column}')
                print(f'row_heights_per_column_no_wrap = {row_heights_per_column_no_wrap}')
            handle_merged_column_widths(col_widths)
            handle_merged_column_widths(col_widths_no_wrap)
            handle_merged_row_heights(row_heights)
            handle_merged_row_heights(row_heights_per_column)
            handle_merged_row_heights(row_heights_per_column_no_wrap)
            if TRACE:
                print("Values after handling merged cells:")

        # Try not to word-wrap as many columns as possible, lopping them off at the end if they are too wide
        width = 0
        for wid in col_widths_no_wrap.values():
            width += wid

        if TRACE:
            print(f'col_widths                     = {col_widths}')
            print(f'col_widths_no_wrap             = {col_widths_no_wrap}')
            print(f'row_heights                    = {row_heights}')
            print(f'row_heights_per_column         = {row_heights_per_column}')
            print(f'row_heights_per_column_no_wrap = {row_heights_per_column_no_wrap}')
            print(f'fully_merged_rows              = {fully_merged_rows}')

        # FIXME: if the height of a cell is greater than the MAX_CELL_HEIGHT_PT, then try
        # widening that cell in order to fix the text into it better.  Check on the latin
        # text in test case "width.xls".

        max_spreadsheet_width = CSSStyle.px_to_units(CSSStyle.SPREADSHEET_WIDTH_PX)
        while width > max_spreadsheet_width:
            for col in reversed(list(col_widths_no_wrap)):
                wid = col_widths_no_wrap[col]
                if wid > col_widths[col]:
                    width -= (wid - col_widths[col])
                    if TRACE:
                        print(f'Sheet too wide: Removing no_wrap from col {col}')
                    col_widths_no_wrap[col] = col_widths[col]
                    for row in row_heights_per_column_no_wrap:
                        if row in fully_merged_rows:
                            continue
                        if col in row_heights_per_column[row]:
                            row_heights_per_column_no_wrap[row][col] = row_heights_per_column[row][col]
                    break
            else:
                break

        for col, wid in col_widths_no_wrap.items():
            wid = CSSStyle.fixup_excel_width(wid)
            if TRACE:
                print(f'ws.column_dimensions[{col}].width = {wid}')
            if col in col_max_widths:
                max_width = CSSStyle.fixup_excel_width(col_max_widths[col])
                if wid > max_width:
                    wid = max_width
            if wid == 0.0:
                ws.column_dimensions[col].width = 0.0
                ws.column_dimensions[col].hidden = True
            else:
                ws.column_dimensions[col].width = wid

        for row, cols in row_heights_per_column_no_wrap.items():
            height = 0
            if row in fully_merged_rows:
                if TRACE:
                    print(f'Using wrapped height for fully_merged row {row}')
                cols = row_heights_per_column[row]      # Wrap it if it's super-wide
            for h in cols.values():
                height = max(height, h)
            if height != 0:
                row_heights[row] = max(row_heights.get(row, 0), height)

        for row, height in row_heights.items():
            if TRACE:
                print(f'ws.row_dimensions[{row}].height = {height}')
            if height == 0.0:
                ws.row_dimensions[row].height = 0.0
                ws.row_dimensions[row].hidden = True
            else:
                ws.row_dimensions[row].height = min(CSSStyle.MAX_CELL_HEIGHT_PT, max(CSSStyle.MIN_CELL_HEIGHT_PT, height))

        # If we had multiple images per cell, place them now that we know the column widths
        for image_data in extra_images:
            row, col, wid, img = image_data
            offset = 0
            cl = get_column_letter(col)
            anchor = f'{cl}{row}'
            for c in range(col, global_n_cols+1):
                cl = get_column_letter(c)
                w = CSSStyle.units_to_px(col_widths_no_wrap.get(cl, CSSStyle.px_to_units(CSSStyle.DEFAULT_CELL_WIDTH_PX)))
                offset += w
                anchor = f'{cl}{row}'
                if anchor not in has_image and offset >= wid:
                    break

            img.anchor = anchor
            ws.add_image(img)

        if filename:
            wb.save(filename=filename)
            return filename
        return wb
