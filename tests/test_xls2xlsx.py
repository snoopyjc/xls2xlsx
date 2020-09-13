#!/usr/bin/env python

"""Tests for `xls2xlsx` package."""

import pytest
import colorsys
import os
import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import RGB, COLOR_INDEX
from openpyxl.styles.borders import Border

from xls2xlsx import XLS2XLSX
from xls2xlsx.utils import perform_number_format

DEFAULT_ROW_HEIGHT=15   # In Excel units
DEFAULT_COLUMN_WIDTH=8.43
MAX_WIDTH_HEIGHT_DIFF=25/100        # 25% off to fail

URL='http://www.snoopyjc.org/pytest/xls2xlsx'
REMOTE_TESTS={'Styles3.xls', 'Styles4.mht', 'Styles5.htm'}
TESTS='tests'
INPUTS=os.path.join(TESTS, 'inputs')
OUTPUTS=os.path.join(TESTS, 'outputs')
EXPECTED=os.path.join(TESTS, 'expected')
TESTS=set([os.path.join(INPUTS, f) for f in os.listdir(INPUTS) if '_' not in f and '.' in f and f[0] != '.' and
    not f.endswith('~') and not f.endswith('.out') and not f.endswith('.log') and os.path.isfile(os.path.join(INPUTS, f))])

EXPECTED_FAILURES={'Styles1.xls': {'Sheet1': {'all': 'image', 'A13': 'alignment'}},    # xlrd doesn't pass thru images
        'Styles3.xls': {'Sheet1': {'all': 'image', 'A13': 'alignment'}}}    # And doesn't auto-right-align percentages (the HTML version does)
# Same thing: HTML version sets align=right:
EXPECTED_FAILURES.update({'Formatf.xls': {'Blätt1': dict(), 'ÖÄÜ': dict(), 'Blätt3': dict()}})
for row in range(1, 11):
    EXPECTED_FAILURES['Formatf.xls']['Blätt1'][f'B{row}'] = 'alignment'

EXPECTED_FAILURES['Formatf.xls']['ÖÄÜ']['A1'] = 'alignment'

# HTML version goes from A1:I17 vs xls from A1:A12 because of chart, also xls version skips the chart totally!
EXPECTED_FAILURES['Formatf.xls']['Blätt3']['all'] = 'dimension,image' 
for row in range(1, 13):
    EXPECTED_FAILURES['Formatf.xls']['Blätt3'][f'A{row}'] = 'alignment' # general vs right in the HTML version

for row in range(21, 41):           # Again, the auto-right-align is applied to these cells
    EXPECTED_FAILURES['Styles1.xls']['Sheet1'][f'B{row}'] = 'alignment'
    EXPECTED_FAILURES['Styles3.xls']['Sheet1'][f'B{row}'] = 'alignment'
EXPECTED_FAILURES['Styles1.xls']['Sheet1']['B35'] += ',font'  # HTML version pre-colors the font
EXPECTED_FAILURES['Styles3.xls']['Sheet1']['B35'] += ',font'  # HTML version pre-colors the font
for col in range(ord('B'), ord('J')):               # HTML version centers the errors
    c = chr(col)
    EXPECTED_FAILURES['Styles1.xls']['Sheet1'][f'{c}43'] = 'alignment'
    EXPECTED_FAILURES['Styles3.xls']['Sheet1'][f'{c}43'] = 'alignment'
EXPECTED_FAILURES['Styles1.xls']['Sheet1']['A49'] = 'value'   # HTML version chops 1 letter off
EXPECTED_FAILURES['Styles3.xls']['Sheet1']['A49'] = 'value'   # HTML version chops 1 letter off
EXPECTED_FAILURES['Styles1.xls']['Sheet1']['A55'] = 'value'   # HTML version chops 1 letter off
EXPECTED_FAILURES['Styles3.xls']['Sheet1']['A55'] = 'value'   # HTML version chops 1 letter off
EXPECTED_FAILURES['Styles1.xls']['Sheet1']['B40'] += ',font'     # Colors are off!
EXPECTED_FAILURES['Styles3.xls']['Sheet1']['B40'] += ',font'     # Colors are off!
EXPECTED_FAILURES['Styles1.xls']['Sheet2'] = {'all': 'dimension'}   # HTML version has 2 columns
EXPECTED_FAILURES['Styles3.xls']['Sheet2'] = {'all': 'dimension'}   # HTML version has 2 columns
EXPECTED_FAILURES['Styles1.xls']['Sheet2']['B1'] = 'font'
EXPECTED_FAILURES['Styles3.xls']['Sheet2']['B1'] = 'font'
#EXPECTED_FAILURES['Styles1.xls']['Sheet3'] = {'A1': 'font'}     # Using a themed font
#EXPECTED_FAILURES['Styles3.xls']['Sheet3'] = {'A1': 'font'}

EXPECTED_FAILURES.update({'Extras1.xls': {'Sheet1': {'all': 'image'}}}) # xlrd eats images
EXPECTED_FAILURES.update({'Extras2.xls': {'Sheet1': {'all': 'image'}}}) # Real xlsx only shows 1 image, the rest are other things (e.g. text box)
EXPECTED_FAILURES['timedelta2.mht'] = {'Sheet1': {'all': 'dimension'}}  # Extra row in HTML version

for col in ('B', 'C', 'N', 'O'):    # Numbers are right-justified in the HTML, but None or 'general' in the xlsx
    for row in range(12, 17):
        number = f'{col}{row}'
        EXPECTED_FAILURES['Extras2.xls']['Sheet1'][number] = 'alignment'

for pivot in ('M11', 'N11', 'O11', 'M16', 'N16', 'O16'):
    EXPECTED_FAILURES['Extras2.xls']['Sheet1'][pivot] = 'font,fill,border,alignment' # The real xlsx doesn't have it marked bold, fill is not set, and neither is border

for formula in ('B29', 'B30', 'B31', 'B32'):
    EXPECTED_FAILURES['Extras1.xls']['Sheet1'][formula] = 'value,alignment'     # Formula vs value
    EXPECTED_FAILURES['Extras2.xls']['Sheet1'][formula] = 'value,alignment'     # Formula vs value

for conditional in range(28, 33):
    EXPECTED_FAILURES['Extras2.xls']['Sheet1'][f'E{conditional}'] = 'fill,alignment'     # Conditional formatting built-in HTML version
    EXPECTED_FAILURES['Extras2.xls']['Sheet1'][f'F{conditional}'] = 'alignment'          # Numbers are right-justified in the HTML

EXPECTED_FAILURES['Dates.xls'] = {'Sheet1': dict()}
for row in range(2, 34):
    EXPECTED_FAILURES['Dates.xls']['Sheet1'][f'A{row}'] = 'number_format'   # General vs specific date format

for row in range(2, 8):
    EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'A{row}'] = 'alignment'
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'A2'] += 'number_format'
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'A3'] += 'number_format'
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'B2'] = 'number_format'
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'B3'] = 'number_format'
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'B4'] = 'number_format'
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'B6'] = 'value'  # str vs int
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'B7'] = 'number_format'  # General vs Custom
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'A8'] = 'alignment'  # None vs Right
EXPECTED_FAILURES['timedelta2.mht']['Sheet1'][f'B8'] = 'value'  # str vs int

for remote_test in REMOTE_TESTS:
    TESTS.remove(os.path.join(INPUTS, remote_test))
    TESTS.add(f'{URL}/{remote_test}')

color_index = list(COLOR_INDEX)
color_index[64] = '00000000'      # Index 64: System foreground
color_index[65] = '00FFFFFF'      # Index 65: System background

def close_color(rgb1, rgb2):
    """xls colors are mapped using a color palette which doesn't accurately represent the
    colors in the spreadsheet, e.g. the color for a BAD formatted cell is off.  Here we
    see if the color they picked is close enough to the actual color to say we passed the test.
    Note: Somehow Excel is able to figure out the real color that was used in a xls file,
    but xlrd and OpenOffice are not, so it's undocumented!"""
    # See also: https://bz.apache.org/ooo/show_bug.cgi?id=110667

    if rgb1 is None or isinstance(rgb1, RGB):    # None is black
        rgb1 = '0'
    if rgb2 is None or isinstance(rgb2, RGB):
        rgb2 = '0'
    rgb1 = int(rgb1, 16) & 0xffffff
    rgb2 = int(rgb2, 16) & 0xffffff
    if rgb1 == rgb2:
        return True
    yiq1 = colorsys.rgb_to_yiq((rgb1>>24) & 0xff, (rgb1>>16) & 0xff, rgb1 & 0xff)
    yiq2 = colorsys.rgb_to_yiq((rgb2>>24) & 0xff, (rgb2>>16) & 0xff, rgb2 & 0xff)
    MAX_LUMA_DISTANCE=5
    MAX_IN_PHASE_DISTANCE=32
    MAX_QUADRATURE_DISTANCE=55
    if abs(yiq1[0]-yiq2[0]) < MAX_LUMA_DISTANCE and \
       abs(yiq1[1]-yiq2[1]) < MAX_IN_PHASE_DISTANCE and \
       abs(yiq1[2]-yiq2[2]) < MAX_QUADRATURE_DISTANCE:
       return True
    print(f'rgb1 = {rgb1:06x}, yiq1 = {yiq1}, rgb2 = {rgb2:06x}, yiq2 = {yiq2}')
    return False

def eq(o1, o2):
    if o1 is None or o2 is None:
        return o1 == o2
    if type(o1) is not type(o2):
        return False
    if isinstance(o1, (list, tuple, set, dict, str, bool, int, float)):
        return o1 == o2

    ignore_bgColor = False
    if hasattr(o1, 'style') and hasattr(o2, 'style'):
        if o1.style is None and o2.style is None:
            return True     # For Border Side objects, if there is no style, ignore the color
    if hasattr(o1, 'patternType') and hasattr(o2, 'patternType'):
        if o1.patternType is None and o2.patternType is None:
            return True     # For PatternFill objects, if there is no pattern, ignore both colors
        if o1.patternType == 'solid' and o2.patternType == 'solid':
            ignore_bgColor = True   # If it's a solid fill, we ignore the background color
    if hasattr(o1, 'type') and hasattr(o2, 'type'):
        if o1.type == 'theme' and o2.type == 'rgb' and o1.theme == 1 and o2.rgb == '00000000' and o1.tint == 0.0 and o2.tint == 0.0:
            return True     # Theme color 1 == black
        if o2.type == 'theme' and o1.type == 'rgb' and o2.theme == 1 and o1.rgb == '00000000' and o1.tint == 0.0 and o2.tint == 0.0:
            return True     # Theme color 1 == black
        if o1.type == 'indexed':
            o1 = copy.copy(o1)
            o1.type = 'rgb'
            o1.rgb = color_index[o1.indexed]
            #delattr(o1, 'indexed')
        if o2.type == 'indexed':
            o2 = copy.copy(o2)
            o2.type = 'rgb'
            o2.rgb = color_index[o2.indexed]
            #delattr(o2, 'indexed')
    for v in vars(o1):
        if v.startswith('_'):
            continue
        if v == 'bgColor' and ignore_bgColor:
            continue
        if v == 'family':       # Font family - let it slide if everything else is the same!
            continue
        elif v == 'scheme':     # Color scheme - ignore this
            continue
        elif v == 'indexed':    # Color index (handled above, but not able to delete)
            continue
        elif v == 'color':
            # None color is black
            if o1.color is None and o2.color is None:
                continue
            if o1.color is None and ((o2.color.type == 'theme' and o2.color.theme == 1) or (o2.color.type == 'rgb' and o2.color.rgb == '00000000')):
                continue
            if o2.color is None and ((o1.color.type == 'theme' and o1.color.theme == 1) or (o1.color.type == 'rgb' and o1.color.rgb == '00000000')):
                continue
        elif v == 'rgb':        # Sometimes they are different in the top 8 bits (alpha)!
            if not close_color(o1.rgb, o2.rgb):
                return False
            continue
        elif v == 'strike':     # Sometimes they are False vs None
            if bool(o1.strike) != bool(o2.strike):
                return False
            continue
        elif v == 'horizontal':     # default Alignment
            if o1.horizontal is None and o2.horizontal == 'general':
                continue
            elif o1.horizontal == 'general' and o2.horizontal is None:
                continue
        elif v == 'vertical':       # default Alignment
            if o1.vertical is None and o2.vertical == 'bottom':
                continue
            elif o1.vertical == 'bottom' and o2.vertical is None:
                continue
        if not eq(getattr(o1, v), getattr(o2, v)):
            return False
    return True

def eq_nbs(s1, s2):     # No backslashes
    if s1 is None or s2 is None:
        return s1 == s2
    return s1.replace('\\', '') == s2.replace('\\', '')

def or_border(ws1, ws2, row, col):
    """Sometimes the border's don't quite match as, for example, the left border is optional
    if the right border of the prior cell is set - so account for that by copying the appropriate borders
    if they are set."""
    ws1_cell = ws1.cell(row, col)
    ws2_cell = ws2.cell(row, col)
    if row != ws1.min_row:
        if not ws1_cell.border.top.style and ws2_cell.border.top.style and \
                ws1.cell(row-1, col).border.bottom.style:
            ws1_cell.border = Border(top=ws1.cell(row-1, col).border.bottom, left=ws1_cell.border.left, right=ws1_cell.border.right, bottom=ws1_cell.border.bottom)
        elif not ws2_cell.border.top.style and ws1_cell.border.top.style and \
                ws2.cell(row-1, col).border.bottom.style:
            ws2_cell.border = Border(top=ws2.cell(row-1, col).border.bottom, left=ws2_cell.border.left, right=ws2_cell.border.right, bottom=ws2_cell.border.bottom)
    if row != ws1.max_row:
        if not ws1_cell.border.bottom.style and ws2_cell.border.bottom.style and \
                ws1.cell(row+1, col).border.top.style:
            ws1_cell.border = Border(bottom=ws1.cell(row+1, col).border.top, left=ws1_cell.border.left, right=ws1_cell.border.right, top=ws1_cell.border.top)
        elif not ws2_cell.border.bottom.style and ws1_cell.border.bottom.style and \
                ws2.cell(row+1, col).border.top.style:
            ws2_cell.border = Border(bottom=ws2.cell(row+1, col).border.top, left=ws2_cell.border.left, right=ws2_cell.border.right, top=ws2_cell.border.top)
    if col != ws1.min_column:
        if not ws1_cell.border.left.style and ws2_cell.border.left.style and \
                ws1.cell(row, col-1).border.right.style:
            ws1_cell.border = Border(left=ws1.cell(row, col-1).border.right, right=ws1_cell.border.right, bottom=ws1_cell.border.bottom, top=ws1_cell.border.top)
        elif not ws2_cell.border.left.style and ws1_cell.border.left.style and \
                ws2.cell(row, col-1).border.right.style:
            ws2_cell.border = Border(left=ws2.cell(row, col-1).border.right, right=ws2_cell.border.right, bottom=ws2_cell.border.bottom, top=ws2_cell.border.top)
    if col != ws1.max_column:
        if not ws1_cell.border.right.style and ws2_cell.border.right.style and \
                ws1.cell(row, col+1).border.left.style:
            ws1_cell.border = Border(right=ws1.cell(row, col+1).border.left, left=ws1_cell.border.left, bottom=ws1_cell.border.bottom, top=ws1_cell.border.top)
        elif not ws2_cell.border.right.style and ws1_cell.border.right.style and \
                ws2.cell(row, col+1).border.left.style:
            ws2_cell.border = Border(right=ws2.cell(row, col+1).border.left, left=ws2_cell.border.left, bottom=ws2_cell.border.bottom, top=ws2_cell.border.top)

best_diff = 0.0
def close_value(v1, v2, default):
    global best_diff
    v1 = v1 or default
    v2 = v2 or default
    diff = 1.0 - (min(v1, v2) / max(v1, v2, 0.01))  # The 0.01 is just so we don't divide by 0
    if diff > MAX_WIDTH_HEIGHT_DIFF:
        return False
    return True




@pytest.mark.parametrize('xls', list(TESTS))
def test_one_xls(xls):
    basename = os.path.split(xls)[-1]
    fn, ext = os.path.splitext(basename)
    expected = os.path.join(EXPECTED, fn) + '.xlsx'
    output = os.path.join(OUTPUTS, fn) + '.xlsx'
    inp = xls
    if fn.startswith('Styles'):
        expected = os.path.join(EXPECTED, 'Styles.xlsx')
    elif fn.startswith('Format'):
        expected = os.path.join(EXPECTED, 'Formate.xlsx')
    elif fn.startswith('Extras'):
        expected = os.path.join(EXPECTED, 'Extras.xlsx')
    elif fn.startswith('timedelta'):
        expected = os.path.join(EXPECTED, 'timedelta.xlsx')

    expected_failures = EXPECTED_FAILURES.get(basename, dict())

    x2x = XLS2XLSX(inp)
    x2x.to_xlsx(filename=output)
    wb_exp = load_workbook(filename=expected)
    wb_out = load_workbook(filename=output)

    for i, ws_exp in enumerate(wb_exp.worksheets):
        ws_out = wb_out.worksheets[i]
        assert ws_exp.title == ws_out.title
        expected_sheet_failures = expected_failures.get(ws_exp.title, dict())
        esfa = expected_sheet_failures.get('all', '')
        if 'dimension' not in esfa:
            assert ws_exp.calculate_dimension() == ws_out.calculate_dimension()
        exp_merged = [str(mc) for mc in ws_exp.merged_cells.ranges]
        exp_merged.sort()
        out_merged = [str(mc) for mc in ws_out.merged_cells.ranges]
        out_merged.sort()
        assert exp_merged == out_merged
        images = ws_exp._images
        if images is None:
            assert ws_out._images is None
        elif 'image' not in esfa:
            assert len(images) == len(ws_out._images)

        if 'height' not in esfa:
            for row in range(ws_exp.min_row, ws_exp.max_row+1):
                esf = expected_sheet_failures.get(row, '')
                if 'height' not in esf:
                    exp_rd = ws_exp.row_dimensions[row]
                    out_rd = ws_out.row_dimensions[row]
                    assert exp_rd.hidden == out_rd.hidden
                    assert close_value(exp_rd.height, out_rd.height, DEFAULT_ROW_HEIGHT)
        if 'width' not in esfa:
            for col in range(ws_exp.min_column, ws_exp.max_column+1):
                cl = get_column_letter(col)
                esf = expected_sheet_failures.get(cl, '')
                if 'width' not in esf:
                    exp_cd = ws_exp.column_dimensions[cl]
                    out_cd = ws_out.column_dimensions[cl]
                    assert exp_cd.hidden == out_cd.hidden
                    assert close_value(exp_cd.width, out_cd.width, DEFAULT_COLUMN_WIDTH)

        for row in range(ws_exp.min_row, ws_exp.max_row+1):
            for col in range(ws_exp.min_column, ws_exp.max_column+1):
                cl = get_column_letter(col) + str(row)
                esf = expected_sheet_failures.get(cl, '')
                exp_cell = ws_exp.cell(row, col)
                out_cell = ws_out.cell(row, col)
                # We include the (row, col) to let the user know what cell is failing
                if 'value' not in esf:
                    assert (row, col) == (row, col) and perform_number_format(exp_cell.value, exp_cell.number_format) == perform_number_format(out_cell.value, out_cell.number_format)
                if 'font' not in esf and exp_cell.value is not None:
                    assert (row, col) == (row, col) and eq(exp_cell.font, out_cell.font)
                if 'fill' not in esf:
                    assert (row, col) == (row, col) and eq(exp_cell.fill, out_cell.fill)
                if 'border' not in esf:
                    or_border(ws_exp, ws_out, row, col)
                    assert (row, col) == (row, col) and eq(exp_cell.border, out_cell.border)
                if 'alignment' not in esf and exp_cell.value is not None:
                    assert (row, col) == (row, col) and eq(exp_cell.alignment, out_cell.alignment)
                if 'number_format' not in esf:
                    assert (row, col) == (row, col) and eq_nbs(exp_cell.number_format, out_cell.number_format)
                if 'hyperlink' not in esf:
                    assert (row, col) == (row, col) and eq(exp_cell.hyperlink, out_cell.hyperlink)
                if 'comment' not in esf:
                    assert (row, col) == (row, col) and eq(exp_cell.comment, out_cell.comment)


def test_cleanup():
    """Make sure we clean up after ourself"""
    assert not os.path.exists('http')
    files = os.listdir(INPUTS)
    for f in files:
        if os.path.isdir(os.path.join(INPUTS, f)):
            if f.endswith('_files'):
                continue
            assert f is None            # failed!
